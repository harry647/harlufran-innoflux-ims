# lab_logic.py

import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox
import csv
from plyer import notification
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl
from openpyxl.styles import Font
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate
from reportlab.lib import colors
from datetime import datetime, timedelta
import webbrowser
from constraint import Problem, AllDifferentConstraint
from scipy.optimize import minimize
import statsmodels.api as sm
import numpy as np
from chemistry_practical import SolutionCalculatorApp
import logging
import os


log_dir = os.path.join(os.getenv("APPDATA"), "InventoryManagementSystem", "logs")
os.makedirs(log_dir, exist_ok=True)  # Create directory if it doesn’t exist
log_file = os.path.join(log_dir, "inventory_app.log")

logging.basicConfig(
    filename=log_file,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class LabLogic:
    def __init__(self, app, lab_tab):
        self.app = app
        self.lab_tab = lab_tab
        self.section = "lab"
        self.lab_tree = None  # Initialize here
        self.broken_tree = None # Initialize here
        self.new_qty_entries = {}
        self.chem_prac = None
        self.plan_window = None
        self.item_scrollable_frame = None
        self.plan_items_text = None
        self.plan_result = None
        self.plan_subject = None
        self.plan_students = None
        self.plan_groups = None
        self.plan_date = None
        self.plan_day = None
        self.plan_time = None
        self.plan_duration = None
        self.item_canvas = None
        self.item_frame = None
        self.item_scrollbar = None

        self.setup_lab_tab()
        self.load_items(self.lab_tree)
        self.load_broken_items()
        self.check_reorder()
        self.check_expiry()

    def setup_lab_tab(self):
        # Main Container Frame for Input and Broken Item Reporting
        self.top_frame = ttk.Frame(self.lab_tab)  # Store as instance variable for toggling
        self.top_frame.pack(fill="x", padx=10, pady=5)
        self.top_frame.columnconfigure(0, weight=1) # Input Frame
        self.top_frame.columnconfigure(1, weight=1) # Broken Frame
        self.top_frame.columnconfigure(2, weight=1) # Action Frame

        # Add/Edit Item Frame (Left Side)
        input_frame = ttk.LabelFrame(self.top_frame, text="Add/Edit Item")
        input_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        input_frame.columnconfigure(0, weight=0)
        input_frame.columnconfigure(1, weight=1)

        ttk.Label(input_frame, text="Name:").grid(row=0, column=0, padx=5, pady=2, sticky="w")
        self.lab_name = ttk.Entry(input_frame)
        self.lab_name.grid(row=0, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(input_frame, text="Category:").grid(row=1, column=0, padx=5, pady=2, sticky="w")
        self.lab_category = ttk.Combobox(input_frame, values=self.get_categories())
        self.lab_category.grid(row=1, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(input_frame, text="Subject:").grid(row=2, column=0, padx=5, pady=2, sticky="w")
        self.lab_subject = ttk.Combobox(input_frame, values=["Physics", "Chemistry", "Biology"])
        self.lab_subject.grid(row=2, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(input_frame, text="Quantity (e.g., '10 mL'):").grid(row=3, column=0, padx=5, pady=2, sticky="w")
        self.lab_quantity = ttk.Entry(input_frame)
        self.lab_quantity.grid(row=3, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(input_frame, text="Reorder Level (e.g., '2 mL'):").grid(row=4, column=0, padx=5, pady=2, sticky="w")
        self.lab_reorder = ttk.Entry(input_frame)
        self.lab_reorder.grid(row=4, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(input_frame, text="Expiry Date (YYYY-MM-DD):").grid(row=5, column=0, padx=5, pady=2, sticky="w")
        self.lab_expiry = ttk.Entry(input_frame)
        self.lab_expiry.grid(row=5, column=1, padx=5, pady=2, sticky="ew")

        ttk.Button(input_frame, text="Add Item", command=self.add_lab_item).grid(row=6, column=0, padx=5, pady=5, sticky="ew")
        ttk.Button(input_frame, text="Update Items", command=self.update_lab_items).grid(row=6, column=1, padx=5, pady=5, sticky="ew")

        # Report Broken Item Frame (Right Side)
        broken_frame = ttk.LabelFrame(self.top_frame, text="Report Broken Item")
        broken_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
        broken_frame.columnconfigure(0, weight=0)
        broken_frame.columnconfigure(1, weight=1)

        ttk.Label(broken_frame, text="Student ID:").grid(row=0, column=0, padx=5, pady=2, sticky="w")
        self.broken_student_id = ttk.Entry(broken_frame)
        self.broken_student_id.grid(row=0, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(broken_frame, text="Student Name:").grid(row=1, column=0, padx=5, pady=2, sticky="w")
        self.broken_student_name = ttk.Entry(broken_frame)
        self.broken_student_name.grid(row=1, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(broken_frame, text="Item Name:").grid(row=2, column=0, padx=5, pady=2, sticky="w")
        self.broken_item = ttk.Combobox(broken_frame, values=self.get_item_names())
        self.broken_item.grid(row=2, column=1, padx=5, pady=2, sticky="ew")

        ttk.Button(broken_frame, text="Report Broken", command=self.report_broken_item).grid(row=3, column=1, padx=5, pady=5, sticky="ew")

        # Search Frame
        self.search_frame = ttk.Frame(self.lab_tab)  # Store as instance variable
        self.search_frame.pack(fill="x", padx=10, pady=5)
        self.search_frame.columnconfigure(2, weight=1)  # Adjusted for new widget

        # Filter selection
        ttk.Label(self.search_frame, text="Search by:").pack(side="left", padx=(0, 5))
        self.search_filter = ttk.Combobox(self.search_frame, 
                                        values=["Name", "Category", "ID", "Quantity", "Subject"], 
                                        state="readonly")
        self.search_filter.set("Name")  # Default filter
        self.search_filter.pack(side="left", padx=5)

        # Search entry
        ttk.Label(self.search_frame, text="Search Items:").pack(side="left")
        self.lab_search = ttk.Entry(self.search_frame)
        self.lab_search.pack(side="left", padx=5, fill="x", expand=True)
        self.lab_search.bind("<KeyRelease>", lambda e: self.search_items())

        # Treeview for Inventory Display (removed Price column)
        self.lab_tree = ttk.Treeview(self.lab_tab, 
                                   columns=("ID", "Name", "Category", "Subject", "Quantity", 
                                           "Reorder", "Expiry", "Last Updated"), 
                                   show="headings")
        self.lab_tree.heading("ID", text="ID")
        self.lab_tree.heading("Name", text="Name")
        self.lab_tree.heading("Category", text="Category")
        self.lab_tree.heading("Subject", text="Subject")
        self.lab_tree.heading("Quantity", text="Quantity")
        self.lab_tree.heading("Reorder", text="Reorder Level")
        self.lab_tree.heading("Expiry", text="Expiry Date")
        self.lab_tree.heading("Last Updated", text="Last Updated")
        self.lab_tree.pack(fill="both", expand=1, padx=10, pady=5)
        self.lab_tree.bind("<Button-1>", self.select_item)
        self.lab_tree.bind("<Enter>", lambda e: self.app.show_tooltip(e, "Click to edit item"))
        self.lab_tree.bind("<Leave>", self.app.hide_tooltip)
        
        for col in ("ID", "Name", "Category", "Subject", "Quantity", "Reorder", "Expiry", "Last Updated"):
            self.lab_tree.column(col, stretch=tk.YES)

        # Broken Items Treeview (unchanged)
        self.broken_tree = ttk.Treeview(self.lab_tab, 
                                      columns=("ID", "Student ID", "Student Name", "Item Name", 
                                             "Report Date", "Status"), 
                                      show="headings")
        self.broken_tree.heading("ID", text="ID")
        self.broken_tree.heading("Student ID", text="Student ID")
        self.broken_tree.heading("Student Name", text="Student Name")
        self.broken_tree.heading("Item Name", text="Item Name")
        self.broken_tree.heading("Report Date", text="Report Date")
        self.broken_tree.heading("Status", text="Status")
        self.broken_tree.pack(fill="both", expand=1, padx=10, pady=5)
        self.broken_tree.bind("<Button-1>", self.select_broken_item)
        for col in ("ID", "Student ID", "Student Name", "Item Name", "Report Date", "Status"):
            self.broken_tree.column(col, stretch=tk.YES)

        # Action Frame
        action_frame = ttk.LabelFrame(self.top_frame, text="More Actions")
        action_frame.grid(row=0, column=2, padx=5, pady=5, sticky="nsew")

        # Configure column weights for buttons in Action Frame
        for i in range(5):
            action_frame.columnconfigure(i, weight=1)

        # Row 0: Core Actions + Reporting
        ttk.Button(action_frame, text="Delete Item", command=self.delete_lab_item).grid(row=0, column=0, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Clear Student", command=self.clear_broken_item).grid(row=0, column=1, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Plan Practicals", command=self.plan_practicals).grid(row=0, column=2, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Practical Reports", command=self.manage_practical_reports).grid(row=0, column=3, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Audit Report", command=self.audit_report).grid(row=0, column=4, padx=3, pady=5, sticky="ew")

        # Row 1: Export Actions + Chemistry Practical
        ttk.Button(action_frame, text="Export to CSV", command=self.export_to_csv).grid(row=1, column=0, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Export to Excel", command=self.export_to_excel).grid(row=1, column=1, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Export to PDF", command=self.export_to_pdf).grid(row=1, column=2, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Show Graph", command=self.show_graph).grid(row=1, column=3, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Chemistry Practical Planning", command=self.toggle_chem_prac).grid(row=1, column=4, padx=3, pady=5, sticky="ew")

        # Row 2: Add GMP Analysis
        ttk.Button(action_frame, text="GMP Analysis", command=self.gmp_analysis).grid(row=2, column=0, padx=3, pady=5, sticky="ew")

    def toggle_chem_prac(self):
        """Open or close the Chemistry Practical UI as a separate window."""
        if self.chem_prac is None or not self.chem_prac.root.winfo_exists():
            # Create a new Toplevel window for Chemistry Practical
            chem_window = tk.Toplevel(self.app.root)
            chem_window.title("Chemistry Practical Planning")
            chem_window.geometry("950x750")  # Set a reasonable size
            self.chem_prac = SolutionCalculatorApp(chem_window)
            self.chem_prac.frame.pack(fill="both", expand=True)
            # Ensure closing the window returns to LabLogic
            chem_window.protocol("WM_DELETE_WINDOW", self.close_chem_prac)
        else:
            # Close the existing window
            self.close_chem_prac()

    def close_chem_prac(self):
        """Close the Chemistry Practical window and reset the instance."""
        if self.chem_prac and self.chem_prac.root.winfo_exists():
            self.chem_prac.root.destroy()
        self.chem_prac = None

        # Ensure LabLogic UI remains visible (optional, if needed)
        self.top_frame.pack(fill="x", padx=10, pady=5)
        self.search_frame.pack(fill="x", padx=10, pady=5)
        self.lab_tree.pack(fill="both", expand=1, padx=10, pady=5)
        self.broken_tree.pack(fill="both", expand=1, padx=10, pady=5)

    def plan_practicals(self):
        self.plan_window = tk.Toplevel(self.app.root)
        self.plan_window.title("Plan Laboratory Practicals")
        self.plan_window.geometry("950x750")  # Adjusted initial size

        # Configure grid weights for better resizing
        self.plan_window.columnconfigure(1, weight=1)
        self.plan_window.columnconfigure(2, weight=1)
        self.plan_window.rowconfigure(12, weight=1)
        self.plan_window.rowconfigure(15, weight=1)

        # Core planning fields - Organized in a frame
        core_frame = ttk.Frame(self.plan_window)
        core_frame.grid(row=0, column=0, columnspan=2, padx=10, pady=10, sticky="ew")
        core_frame.columnconfigure(1, weight=1)

        ttk.Label(core_frame, text="Subject:", width=15, anchor="w").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.plan_subject = ttk.Combobox(core_frame, values=["Physics", "Chemistry", "Biology"])
        self.plan_subject.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        self.plan_subject.bind("<<ComboboxSelected>>", self.update_item_list)

        ttk.Label(core_frame, text="Total Students:", width=15, anchor="w").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.plan_students = ttk.Entry(core_frame)
        self.plan_students.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(core_frame, text="Number of Groups:", width=15, anchor="w").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.plan_groups = ttk.Entry(core_frame)
        self.plan_groups.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
        

        ttk.Label(core_frame, text="Date (YYYY-MM-DD):", width=15, anchor="w").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.plan_date = ttk.Entry(core_frame)
        self.plan_date.grid(row=3, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(core_frame, text="Day:", width=15, anchor="w").grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.plan_day = ttk.Combobox(core_frame, values=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"])
        self.plan_day.grid(row=4, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(core_frame, text="Start Time (HH:MM):", width=15, anchor="w").grid(row=5, column=0, padx=5, pady=5, sticky="w")
        self.plan_time = ttk.Entry(core_frame)
        self.plan_time.grid(row=5, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(core_frame, text="Duration (hours):", width=15, anchor="w").grid(row=6, column=0, padx=5, pady=5, sticky="w")
        self.plan_duration = ttk.Entry(core_frame)
        self.plan_duration.grid(row=6, column=1, padx=5, pady=5, sticky="ew")

        # Item selection - Organized in a frame
        item_group = ttk.LabelFrame(self.plan_window, text="Available Items")
        item_group.grid(row=0, column=2, rowspan=13, padx=10, pady=10, sticky="nsew")
        item_group.columnconfigure(0, weight=1)
        item_group.rowconfigure(1, weight=1)

        self.item_frame = ttk.Frame(item_group)
        self.item_frame.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
        self.item_canvas = tk.Canvas(self.item_frame)
        self.item_scrollbar = ttk.Scrollbar(self.item_frame, orient="vertical", command=self.item_canvas.yview)
        self.item_scrollable_frame = ttk.Frame(self.item_canvas)

        self.item_scrollable_frame.bind("<Configure>", lambda e: self.item_canvas.configure(scrollregion=self.item_canvas.bbox("all")))
        self.item_canvas.create_window((0, 0), window=self.item_scrollable_frame, anchor="nw")
        self.item_canvas.configure(yscrollcommand=self.item_scrollbar.set)

        self.item_canvas.pack(side="left", fill="both", expand=True)
        self.item_scrollbar.pack(side="right", fill="y")

        ttk.Label(self.plan_window, text="Selected Items (Item:Quantity):", anchor="w").grid(row=13, column=0, padx=10, pady=5, sticky="w")
        self.plan_items_text = tk.Text(self.plan_window, height=4, width=50)
        self.plan_items_text.grid(row=13, column=1, columnspan=2, padx=10, pady=5, sticky="ew")

        # Buttons - Organized in a frame
        button_frame = ttk.Frame(self.plan_window)
        button_frame.grid(row=14, column=0, columnspan=3, padx=10, pady=10, sticky="ew")
        button_frame.columnconfigure(0, weight=1)
        button_frame.columnconfigure(1, weight=1)
        button_frame.columnconfigure(2, weight=1)

        ttk.Button(button_frame, text="Suggest Schedule", command=self.suggest_optimal_schedule).grid(row=0, column=0, pady=5, sticky="ew")
        ttk.Button(button_frame, text="Check Availability", command=self.check_practical_items).grid(row=0, column=1, pady=5, sticky="ew")
        ttk.Button(button_frame, text="Print Plan", command=self.print_plan).grid(row=0, column=2, pady=5, sticky="ew")

        self.plan_result = tk.Text(self.plan_window, height=10, width=80, wrap="word")
        self.plan_result.grid(row=15, column=0, columnspan=3, padx=10, pady=10, sticky="nsew")

        self.update_item_list(None)



    def get_chemicals(self, is_stock=True):
        """Fetch stock chemicals or diluents from items table."""
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        pattern = '%Stock%' if is_stock else '%Water%'  # Default to water for diluents
        c.execute("SELECT name FROM items WHERE section=? AND name LIKE ?", (self.section, pattern))
        items = [row[0] for row in c.fetchall()]
        conn.close()
        return items if items else ["No items found"]

    def update_molarity(self, event):
        """Update stock molarity field when a chemical is selected."""
        chemical = self.chemical_select.get()
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT molarity FROM items WHERE section=? AND name=?", (self.section, chemical))
        row = c.fetchone()
        conn.close()
    
        self.stock_molarity.config(state="normal")
        self.stock_molarity.delete(0, tk.END)
        self.stock_molarity.insert(0, str(row[0]) if row and row[0] else "N/A")
        self.stock_molarity.config(state="readonly")


    def suggest_optimal_schedule(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
    
        # Fetch existing practical times and required items
        c.execute("SELECT time FROM practical_reports WHERE time > ?", (datetime.now().strftime("%Y-%m-%d %H:%M"),))
        booked_times = [row[0] for row in c.fetchall()]
    
        # Define available time slots (e.g., next 5 days, 9 AM - 5 PM, 2-hour slots)
        start_date = datetime.now() + timedelta(days=1)
        time_slots = []
        for day in range(5):
            current = start_date + timedelta(days=day)
            for hour in range(9, 17, 2):  # 9 AM to 5 PM
                slot = current.replace(hour=hour, minute=0, second=0, microsecond=0)
                time_slots.append(slot.strftime("%Y-%m-%d %H:%M"))
    
        # Set up constraint problem
        problem = Problem()
        problem.addVariable("time", [t for t in time_slots if t not in booked_times])
    
        # Add constraint: ensure sufficient item stock
        items_text = self.plan_items_text.get("1.0", tk.END).strip().split("\n")
        required_items = {line.split(":")[0].strip(): line.split(":")[1].strip() for line in items_text if ":" in line}
        for item_name, req_qty in required_items.items():
            c.execute("SELECT quantity FROM items WHERE section=? AND name=?", (self.section, item_name))
            row = c.fetchone()
            if row:
                avail_qty = row[0]
                req_num, req_unit = self.app.parse_quantity(req_qty)
                avail_num, avail_unit = self.app.parse_quantity(avail_qty)
                avail_converted = self.app.convert_units(avail_num, avail_unit, req_unit)
                problem.addConstraint(lambda t, avail=avail_converted, req=req_num: avail >= req, ["time"])
    
        solution = problem.getSolution()
        conn.close()
    
        if solution:
            suggested_time = datetime.strptime(solution["time"], "%Y-%m-%d %H:%M")
            self.plan_date.delete(0, tk.END)
            self.plan_date.insert(0, suggested_time.strftime("%Y-%m-%d"))
            self.plan_time.delete(0, tk.END)
            self.plan_time.insert(0, suggested_time.strftime("%H:%M"))
            self.plan_day.set(suggested_time.strftime("%A"))
        else:
            messagebox.showwarning("Warning", "No available time slot found with sufficient resources.")


    def suggest_practical_time(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT time FROM practical_reports WHERE time > ?", (datetime.now().strftime("%Y-%m-%d %H:%M"),))
        booked_times = [row[0] for row in c.fetchall()]
        conn.close()
    
        suggested_time = datetime.now() + timedelta(days=1)  # Start tomorrow
        while suggested_time.strftime("%Y-%m-%d %H:%M") in booked_times:
            suggested_time += timedelta(hours=2)  # Assume 2-hour slots
        self.plan_date.delete(0, tk.END)
        self.plan_date.insert(0, suggested_time.strftime("%Y-%m-%d"))
        self.plan_time.delete(0, tk.END)
        self.plan_time.insert(0, suggested_time.strftime("%H:%M"))
        self.plan_day.set(suggested_time.strftime("%A"))  # Set day of week


    def update_item_list(self, event):
        for widget in self.item_scrollable_frame.winfo_children():
            widget.destroy()
        
        subject = self.plan_subject.get()
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        if subject:
            c.execute("SELECT name, quantity FROM items WHERE section=? AND subject=?", (self.section, subject))
        else:
            c.execute("SELECT name, quantity FROM items WHERE section=?", (self.section,))
        items = c.fetchall()
        conn.close()
        
        for i, (name, qty) in enumerate(items):
            btn = ttk.Button(self.item_scrollable_frame, text=f"{name} ({qty})", 
                            command=lambda n=name: self.add_item_to_plan(n))
            btn.grid(row=i, column=0, padx=5, pady=2, sticky="w")

    def add_item_to_plan(self, item_name):
        current_text = self.plan_items_text.get("1.0", tk.END).strip()
        qty_prompt = tk.simpledialog.askstring("Quantity", f"Enter quantity for {item_name} (e.g., '5 mL'):")
        if qty_prompt:
            new_line = f"{item_name}:{qty_prompt}"
            if current_text:
                self.plan_items_text.delete("1.0", tk.END)
                self.plan_items_text.insert("1.0", f"{current_text}\n{new_line}")
            else:
                self.plan_items_text.insert("1.0", new_line)

    def check_practical_items(self):
        try:
            total_students = int(self.plan_students.get())
            num_groups = int(self.plan_groups.get())
            if num_groups <= 0 or total_students < num_groups:
                raise ValueError("Invalid number of groups")
            students_per_group = total_students // num_groups
        
            # Use stored preparation if available
            if self.prep_result is None:
                raise ValueError("Run 'Calculate Preparation' first.")
        
            chemical = self.chemical_select.get()
            conn = sqlite3.connect(self.app.get_db_path())
            c = conn.cursor()
            c.execute("SELECT molarity, quantity FROM items WHERE section=? AND name=?", (self.section, chemical))
            chem_row = c.fetchone()
            if not chem_row or chem_row[0] is None:
                raise ValueError(f"No molarity data for {chemical}")
            stock_molarity, stock_qty = chem_row[0], chem_row[1]
        
            
            date = self.plan_date.get()
            day = self.plan_day.get()
            time = self.plan_time.get()
            duration = float(self.plan_duration.get())
        
            result = f"Practical Planning Suggestions for {self.plan_subject.get()}:\n"
            result += f"Date: {date}, Day: {day}, Time: {time}, Duration: {duration} hrs\n"
            result += f"Total Students: {total_students}, Groups: {num_groups}, Students/Group: {students_per_group}\n\n"
        
            
        
            # Inventory check
            result += "Inventory Check:\n"
            stock_num, stock_unit = self.app.parse_quantity(stock_qty)
            stock_available = self.app.convert_units(stock_num, stock_unit, "mL")
            stock_needed = prep['V_stock'] * 1000
            if stock_available < stock_needed:
                result += f"{chemical}: Available {stock_qty}, Needed {stock_needed:.1f} mL, Shortfall: {stock_needed - stock_available:.1f} mL\n"
            else:
                result += f"{chemical}: Available {stock_qty}, Needed {stock_needed:.1f} mL, Sufficient\n"
        
        
            # Other items
            for item_name, req_qty in required_items.items():
                if item_name not in [chemical, diluent]:
                    c.execute("SELECT quantity FROM items WHERE section=? AND name=?", (self.section, item_name))
                    row = c.fetchone()
                    req_num, req_unit = self.app.parse_quantity(req_qty)
                    req_total = req_num * num_groups
                    if row:
                        avail_qty = row[0]
                        avail_num, avail_unit = self.app.parse_quantity(avail_qty)
                        avail_converted = self.app.convert_units(avail_num, avail_unit, req_unit)
                        if avail_converted < req_total:
                            result += f"{item_name}: Available {avail_qty}, Required {req_total} {req_unit}, Shortfall: {req_total - avail_converted:.1f} {req_unit}\n"
                        else:
                            result += f"{item_name}: Available {avail_qty}, Required {req_total} {req_unit}, Sufficient\n"
                    else:
                        result += f"{item_name}: Not found in inventory\n"
        
            conn.close()
        
            self.plan_result.delete("1.0", tk.END)
            self.plan_result.insert("1.0", result)
        
        except ValueError as e:
            messagebox.showerror("Error", str(e) or "Invalid input: Check all fields.")
   
    

    def simulate_scenario(self):
        try:
            original_students = int(self.plan_students.get())
            new_students = tk.simpledialog.askinteger("Simulate", "Enter new total students:", initialvalue=original_students)
            if new_students is None or new_students <= 0:
                return
        
            # Temporarily adjust students and re-run check
            self.plan_students.delete(0, tk.END)
            self.plan_students.insert(0, str(new_students))
            self.check_practical_items()
        
            # Restore original value
            self.plan_students.delete(0, tk.END)
            self.plan_students.insert(0, str(original_students))
            self.plan_result.insert(tk.END, f"\n\nSimulation with {new_students} students completed. Original restored.")
    
        except ValueError:
            messagebox.showerror("Error", "Invalid student number")

    def send_reminder(self, practical_datetime):
        notification.notify(
            title="Practical Reminder",
            message=f"Practical for {self.plan_subject.get()} starts at {practical_datetime.strftime('%Y-%m-%d %H:%M')}",
            timeout=10
        )

    def print_plan(self):
        result = self.plan_result.get("1.0", tk.END).strip()
        if not result:
            messagebox.showwarning("Warning", "No plan to print")
            return
        filename = f"practical_plan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph("Laboratory Practical Plan", styles['Title']))
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        elements.append(Paragraph("<br/><br/>", styles['Normal']))
        elements.append(Paragraph(result.replace("\n", "<br/>"), styles['Normal']))
        doc.build(elements)
        messagebox.showinfo("Success", f"Plan saved as {filename}")
        webbrowser.open(filename)  # Cross-platform printing

    def get_categories(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT DISTINCT category FROM items WHERE section=? AND category IS NOT NULL", (self.section,))
        categories = [row[0] for row in c.fetchall()]
        conn.close()
        return categories or ["General"]

    def get_item_names(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name FROM items WHERE section=?", (self.section,))
        names = [row[0] for row in c.fetchall()]
        conn.close()
        return names

        
    def add_lab_item(self):
        name = self.lab_name.get().strip()
        category = self.lab_category.get().strip() or "General"
        subject = self.lab_subject.get().strip() or "General"
        quantity = self.lab_quantity.get().strip()
        reorder = self.lab_reorder.get().strip() or "0"
        expiry_date = self.lab_expiry.get().strip()
        last_updated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not name or not quantity:
            messagebox.showerror("Error", "Name and Quantity are required!")
            return

        # Check if item exists
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT id FROM items WHERE name=? AND section=?", (name, self.section))
        existing_item = c.fetchone()

        if existing_item:
            response = messagebox.askyesno("Item Exists", f"Item '{name}' already exists in the lab. Would you like to update it instead?")
            conn.close()
            if response:
                self.update_lab_items()
                return
            else:
                messagebox.showinfo("Info", "Add cancelled. Use a different name or update the existing item.")
                return

        # Proceed with adding new item
        price = tk.simpledialog.askfloat("Price", f"Enter price per purchase unit for {name} (Ksh):", minvalue=0.0) or 0.0

        c.execute("""
            INSERT INTO items (name, category, subject, quantity, reorder_level, expiry_date, section, last_updated, purchase_unit) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'unit')
        """, (name, category, subject, quantity, reorder, expiry_date, self.section, last_updated))
        item_id = c.lastrowid

        c.execute("INSERT INTO prices (item_id, price, last_updated) VALUES (?, ?, ?)", (item_id, price, last_updated))
        c.execute("INSERT INTO history (action, item_id, details, timestamp) VALUES (?, ?, ?, ?)",
                  ("Add", item_id, f"Added {name} (Subject: {subject}) with quantity {quantity}, Price Ksh{price:.2f}", last_updated))

        conn.commit()
        conn.close()

        self.load_items(self.lab_tree)
        self.app.load_history()
        self.check_reorder()
        self.check_expiry()
        self.lab_category['values'] = self.get_categories()
        self.clear_lab_entries()

    def update_lab_items(self):
        name = self.lab_name.get().strip()
        quantity_str = self.lab_quantity.get().strip()
        category = self.lab_category.get().strip()
        subject = self.lab_subject.get().strip()
        reorder_str = self.lab_reorder.get().strip()
        expiry_date = self.lab_expiry.get().strip()

        if not name or not quantity_str:
            messagebox.showerror("Error", "Name and Quantity are required to update an item!")
            return

        # Validate quantity
        new_num, new_unit = self.app.parse_quantity(quantity_str)
        if new_num is None:
            messagebox.showerror("Error", "Invalid quantity format (e.g., '10 mL')!")
            return

        # Validate reorder level (optional)
        reorder_num, reorder_unit = self.app.parse_quantity(reorder_str) if reorder_str else (None, None)
        if reorder_str and reorder_num is None:
            messagebox.showerror("Error", "Invalid reorder level format (e.g., '2 mL')!")
            return

        # Validate expiry date (optional)
        if expiry_date:
            try:
                datetime.strptime(expiry_date, "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("Error", "Invalid expiry date format (YYYY-MM-DD)!")
                return

        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()

        # Check if item exists
        c.execute("SELECT id, quantity, purchase_unit FROM items WHERE name=? AND section=?", (name, self.section))
        item = c.fetchone()

        if not item:
            messagebox.showerror("Error", f"Item '{name}' not found in lab section!")
            conn.close()
            return

        item_id, current_qty, purchase_unit = item
        curr_num, curr_unit = self.app.parse_quantity(current_qty)
        if curr_num is None:
            messagebox.showerror("Error", f"Current quantity for {name} is invalid!")
            conn.close()
            return

        # Convert and update quantity
        converted_new_num = self.app.convert_units(new_num, new_unit, curr_unit or purchase_unit)
        updated_num = curr_num + converted_new_num
        updated_qty = f"{updated_num} {curr_unit or purchase_unit}"

        # Update reorder level if provided
        updated_reorder = None
        if reorder_num is not None:
            converted_reorder = self.app.convert_units(reorder_num, reorder_unit, curr_unit or purchase_unit)
            updated_reorder = f"{converted_reorder} {curr_unit or purchase_unit}"

        last_updated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        price = tk.simpledialog.askfloat("Price Update", f"Update price per purchase unit for {name} (Ksh):", minvalue=0.0)

        # Update item in database
        c.execute("UPDATE items SET quantity=?, category=?, subject=?, reorder_level=?, expiry_date=?, last_updated=? WHERE id=?",
                  (updated_qty, category, subject, updated_reorder or reorder_str, expiry_date or None, last_updated, item_id))
        if price is not None:
            c.execute("INSERT OR REPLACE INTO prices (item_id, price, last_updated) VALUES (?, ?, ?)",
                      (item_id, price, last_updated))
        c.execute("INSERT INTO history (action, item_id, details, timestamp) VALUES (?, ?, ?, ?)",
                  ("Update Qty", item_id, f"Added {quantity_str} to {current_qty}, Price Ksh{price or 0:.2f}", last_updated))

        conn.commit()
        conn.close()

        # Refresh UI
        self.load_items(self.lab_tree)
        self.app.load_history()
        messagebox.showinfo("Success", f"Item '{name}' updated successfully!")
        self.clear_lab_entries()

    def clear_lab_entries(self):
        """Clear all input fields in the lab tab."""
        self.lab_name.delete(0, tk.END)
        self.lab_category.set("")
        self.lab_subject.set("")
        self.lab_quantity.delete(0, tk.END)
        self.lab_reorder.delete(0, tk.END)
        self.lab_expiry.delete(0, tk.END)

    def delete_lab_item(self):
        selected = self.lab_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Select an item to delete")
            return
        if messagebox.askyesno("Confirm", "Are you sure you want to delete this item?"):
            item_id = self.lab_tree.item(selected)["values"][0]
            name = self.lab_tree.item(selected)["values"][1]
            conn = sqlite3.connect(self.app.get_db_path())
            c = conn.cursor()
            c.execute("DELETE FROM items WHERE id=?", (item_id,))
            c.execute("INSERT INTO history (action, item_id, details, timestamp) VALUES (?, ?, ?, ?)",
                     ("Delete", item_id, f"Deleted {name}", datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            conn.commit()
            conn.close()
            self.load_items(self.lab_tree)
            self.app.load_history()

    def report_broken_item(self):
        student_id = self.broken_student_id.get()
        student_name = self.broken_student_name.get()
        item_name = self.broken_item.get()
        report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT id, quantity FROM items WHERE section=? AND name=?", (self.section, item_name))
        item = c.fetchone()
        if not item:
            messagebox.showerror("Error", "Item not found")
            conn.close()
            return
        
        item_id, current_quantity = item
        curr_num, curr_unit = self.app.parse_quantity(current_quantity)
        if curr_num < 1:
            messagebox.showerror("Error", "Out of stock")
            conn.close()
            return
        new_quantity = f"{curr_num - 1} {curr_unit}"
        
        c.execute("UPDATE items SET quantity=?, last_updated=? WHERE id=?", (new_quantity, report_date, item_id))
        c.execute("INSERT INTO broken_items (student_id, student_name, item_id, item_name, report_date) VALUES (?, ?, ?, ?, ?)",
                 (student_id, student_name, item_id, item_name, report_date))
        c.execute("INSERT INTO history (action, item_id, details, timestamp) VALUES (?, ?, ?, ?)",
                 ("Break", item_id, f"{item_name} reported broken by {student_name} ({student_id})", report_date))
        conn.commit()
        conn.close()
        
        self.load_items(self.lab_tree)
        self.load_broken_items()
        self.app.load_history()
        self.check_reorder()
        self.check_expiry()
        self.lab_category['values'] = self.get_categories()
        self.clear_broken_entries()

    def select_broken_item(self, event):
        selected = self.broken_tree.selection()
        if selected:
            values = self.broken_tree.item(selected)["values"]
            self.broken_student_id.delete(0, tk.END)
            self.broken_student_id.insert(0, values[1])
            self.broken_student_name.delete(0, tk.END)
            self.broken_student_name.insert(0, values[2])
            self.broken_item.set(values[3])

    def clear_broken_item(self):
        selected = self.broken_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Select a broken item to clear")
            return
        if messagebox.askyesno("Confirm", "Clear this student's broken item record?"):
            broken_id = self.broken_tree.item(selected)["values"][0]
            item_name = self.broken_tree.item(selected)["values"][3]
            conn = sqlite3.connect(self.app.get_db_path())
            c = conn.cursor()
            c.execute("SELECT quantity FROM items WHERE section=? AND name=?", (self.section, item_name))
            current_quantity = c.fetchone()[0]
            curr_num, curr_unit = self.app.parse_quantity(current_quantity)
            new_quantity = f"{curr_num + 1} {curr_unit}"
            
            c.execute("UPDATE broken_items SET status='Cleared' WHERE id=?", (broken_id,))
            c.execute("UPDATE items SET quantity=?, last_updated=? WHERE name=? AND section=?",
                     (new_quantity, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), item_name, self.section))
            c.execute("INSERT INTO history (action, item_id, details, timestamp) VALUES (?, ?, ?, ?)",
                     ("Clear", None, f"Cleared broken {item_name} for student {self.broken_student_id.get()}", datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            conn.commit()
            conn.close()
            self.load_items(self.lab_tree)
            self.load_broken_items()
            self.app.load_history()

    def load_items(self, tree):
        tree.delete(*tree.get_children())
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("""
            SELECT i.id, i.name, i.category, i.subject, i.quantity, i.reorder_level, 
                   i.expiry_date, i.last_updated, p.price, i.purchase_unit 
            FROM items i 
            LEFT JOIN prices p ON i.id = p.item_id 
            WHERE i.section=?
        """, (self.section,))
        for row in c.fetchall():
            item_id, name, category, subject, total_qty, reorder, expiry, last_updated, price, purchase_unit = row  # 10 columns
            if purchase_unit and purchase_unit != 'unit':
                total_num, total_unit = self.app.parse_quantity(total_qty)
                pu_num, _ = self.app.parse_quantity(purchase_unit)
                num_units = total_num / pu_num if pu_num else total_num
                display_qty = f"{num_units:.2f} x {purchase_unit} ({total_qty})"
            else:
                display_qty = total_qty
            tree.insert("", "end", values=(item_id, name, category, subject, display_qty, reorder, expiry, last_updated, f"Ksh{price:.2f}" if price is not None else "N/A"))
        conn.close()

    def load_broken_items(self):
        self.broken_tree.delete(*self.broken_tree.get_children())
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT id, student_id, student_name, item_name, report_date, status FROM broken_items")
        for row in c.fetchall():
            self.broken_tree.insert("", "end", values=row)
        conn.close()

    def search_items(self):
        search_term = self.lab_search.get().lower()
        filter_type = self.search_filter.get()
        self.lab_tree.delete(*self.lab_tree.get_children())
        
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        
        # Define the base query and modify based on filter
        if filter_type == "Name":
            c.execute("SELECT id, name, category, subject, quantity, reorder_level, expiry_date, last_updated FROM items WHERE section=? AND name LIKE ?",
                     (self.section, f"%{search_term}%"))
        elif filter_type == "Category":
            c.execute("SELECT id, name, category, subject, quantity, reorder_level, expiry_date, last_updated FROM items WHERE section=? AND category LIKE ?",
                     (self.section, f"%{search_term}%"))
        elif filter_type == "ID":
            c.execute("SELECT id, name, category, subject, quantity, reorder_level, expiry_date, last_updated FROM items WHERE section=? AND CAST(id AS TEXT) LIKE ?",
                     (self.section, f"%{search_term}%"))
        elif filter_type == "Quantity":
            c.execute("SELECT id, name, category, subject, quantity, reorder_level, expiry_date, last_updated FROM items WHERE section=? AND CAST(quantity AS TEXT) LIKE ?",
                     (self.section, f"%{search_term}%"))
        elif filter_type == "Subject":
            c.execute("SELECT id, name, category, subject, quantity, reorder_level, expiry_date, last_updated FROM items WHERE section=? AND subject LIKE ?",
                     (self.section, f"%{search_term}%"))
        
        # Insert results into Treeview
        for row in c.fetchall():
            self.lab_tree.insert("", "end", values=row)
        
        conn.close()

    def select_item(self, event):
        selected = self.lab_tree.selection()
        if selected:
            values = self.lab_tree.item(selected)["values"]
            self.lab_name.delete(0, tk.END)
            self.lab_name.insert(0, values[1])
            self.lab_category.set(values[2])
            self.lab_subject.set(values[3])
            self.lab_quantity.delete(0, tk.END)
            self.lab_quantity.insert(0, values[4])
            self.lab_reorder.delete(0, tk.END)
            self.lab_reorder.insert(0, values[5])
            self.lab_expiry.delete(0, tk.END)
            self.lab_expiry.insert(0, values[6] if values[6] else "")

    def check_reorder(self):
        if not self.app.notification_enabled:
            return
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name, quantity, reorder_level FROM items WHERE section=?", (self.section,))
        items = c.fetchall()
        conn.close()
    
        for name, qty, reorder in items:
            qty_num, qty_unit = self.app.parse_quantity(qty)
            reorder_num, reorder_unit = self.app.parse_quantity(reorder)
            if reorder_num == 0:  # Skip if no reorder level set
                continue
            qty_converted = self.app.convert_units(qty_num, qty_unit, reorder_unit)
            if qty_converted <= reorder_num:
                notification.notify(
                    title=f"{self.section.capitalize()} Reorder Alert",
                    message=f"{name} is below reorder level ({qty} <= {reorder})",
                    timeout=10
                )

    def check_expiry(self):
        if not self.app.notification_enabled:
            return
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name, expiry_date FROM items WHERE section=? AND expiry_date IS NOT NULL", (self.section,))
        items = c.fetchall()
        conn.close()
        
        now = datetime.now()
        for name, expiry_date in items:
            try:
                expiry = datetime.strptime(expiry_date, "%Y-%m-%d")
                days_left = (expiry - now).days
                if days_left <= 0:
                    notification.notify(
                        title=f"{self.section.capitalize()} Expiry Alert",
                        message=f"{name} has expired on {expiry_date}",
                        timeout=10
                    )
                elif days_left <= 30:
                    notification.notify(
                        title=f"{self.section.capitalize()} Expiry Alert",
                        message=f"{name} expires in {days_left} days ({expiry_date})",
                        timeout=10
                    )
            except ValueError:
                continue
        self.app.root.after(86400000, self.check_expiry)  # Check daily (24 hours in milliseconds)

    def clear_lab_entries(self):
        self.lab_name.delete(0, tk.END)
        self.lab_category.set("")
        self.lab_subject.set("")
        self.lab_quantity.delete(0, tk.END)
        self.lab_reorder.delete(0, tk.END)
        self.lab_expiry.delete(0, tk.END)

    def clear_broken_entries(self):
        self.broken_student_id.delete(0, tk.END)
        self.broken_student_name.delete(0, tk.END)
        self.broken_item.set("")

    def export_to_csv(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name, category, subject, quantity, reorder_level, expiry_date, last_updated FROM items WHERE section=?", (self.section,))
        rows = c.fetchall()
        conn.close()
        
        filename = f"{self.section}_inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        with open(filename, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Name", "Category", "Subject", "Quantity", "Reorder Level", "Expiry Date", "Last Updated"])
            writer.writerows(rows)
        messagebox.showinfo("Success", f"Exported to {filename}")

    def export_to_excel(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name, category, subject, quantity, reorder_level, expiry_date, last_updated FROM items WHERE section=?", (self.section,))
        rows = c.fetchall()
        conn.close()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{self.section.capitalize()} Inventory"
        
        headers = ["Name", "Category", "Subject", "Quantity", "Reorder Level", "Expiry Date", "Last Updated"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        filename = f"{self.section}_inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        messagebox.showinfo("Success", f"Exported to {filename}")

    def export_to_pdf(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name, category, subject, quantity, reorder_level, expiry_date, last_updated FROM items WHERE section=?", (self.section,))
        rows = c.fetchall()
        conn.close()
        
        filename = f"{self.section}_inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []
        data = [["Name", "Category", "Subject", "Quantity", "Reorder Level", "Expiry Date", "Last Updated"]] + list(rows)
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), '#808080'),
            ('TEXTCOLOR', (0, 0), (-1, 0), '#FFFFFF'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), '#F5F5DC'),
            ('GRID', (0, 0), (-1, -1), 1, '#000000'),
        ]))
        elements.append(table)
        
        doc.build(elements)
        messagebox.showinfo("Success", f"Exported to {filename}")

    def show_graph(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name, quantity, reorder_level FROM items WHERE section=?", (self.section,))
        data = c.fetchall()
        conn.close()
        
        names, quantities, reorders = [], [], []
        for name, qty, reorder in data:
            qty_num, _ = self.app.parse_quantity(qty)
            reorder_num, _ = self.app.parse_quantity(reorder)
            names.append(name)
            quantities.append(qty_num)
            reorders.append(reorder_num)
        
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.bar(names, quantities, label="Quantity", color="skyblue")
        ax.plot(names, reorders, "r--", label="Reorder Level")
        ax.set_title(f"{self.section.capitalize()} Inventory Levels")
        ax.set_xlabel("Items")
        ax.set_ylabel("Quantity (numeric only)")
        ax.legend()
        plt.xticks(rotation=45, ha="right")
        
        graph_window = tk.Toplevel(self.app.root)
        graph_window.title(f"{self.section.capitalize()} Inventory Graph")
        canvas = FigureCanvasTkAgg(fig, master=graph_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=1)

    def show_stock_summary(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT COUNT(*), SUM(CAST(SUBSTR(quantity, 1, INSTR(quantity, ' ')-1) AS REAL)), MIN(CAST(SUBSTR(quantity, 1, INSTR(quantity, ' ')-1) AS REAL)), MAX(CAST(SUBSTR(quantity, 1, INSTR(quantity, ' ')-1) AS REAL)) FROM items WHERE section=?", (self.section,))
        total_items, total_qty, min_qty, max_qty = c.fetchone()
        total_items = total_items or 0
        total_qty = total_qty or 0
        min_qty = min_qty or 0
        max_qty = max_qty or 0
        avg_qty = total_qty / total_items if total_items > 0 else 0

        c.execute("SELECT COUNT(*) FROM items WHERE section=? AND CAST(SUBSTR(quantity, 1, INSTR(quantity, ' ')-1) AS REAL) <= CAST(SUBSTR(reorder_level, 1, INSTR(reorder_level, ' ')-1) AS REAL)", (self.section,))
        low_stock = c.fetchone()[0] or 0
        conn.close()

        # Generate report text
        report = f"{self.section.capitalize()} Stock Summary:\n"
        report += f"Total Items: {total_items}\n"
        report += f"Total Quantity (numeric only): {total_qty}\n"
        report += f"Average Quantity (numeric only): {avg_qty:.2f}\n"
        report += f"Minimum Quantity (numeric only): {min_qty}\n"
        report += f"Maximum Quantity (numeric only): {max_qty}\n"
        report += f"Items Below Reorder Level: {low_stock}\n"

        # Create scrollable window with Treeview
        summary_window = tk.Toplevel(self.app.root)
        summary_window.title(f"{self.section.capitalize()} Stock Summary")
        summary_window.geometry("600x400")

        tree = ttk.Treeview(summary_window, columns=("Metric", "Value"), show="headings")
        tree.heading("Metric", text="Metric")
        tree.heading("Value", text="Value")
        tree.column("Metric", width=200)
        tree.column("Value", width=200)

        # Insert data into Treeview
        metrics = [
            ("Total Items", total_items),
            ("Total Quantity (numeric only)", total_qty),
            ("Average Quantity (numeric only)", f"{avg_qty:.2f}"),
            ("Minimum Quantity (numeric only)", min_qty),
            ("Maximum Quantity (numeric only)", max_qty),
            ("Items Below Reorder Level", low_stock)
        ]
        for metric, value in metrics:
            tree.insert("", "end", values=(metric, value))

        # Scrollbar
        scrollbar = ttk.Scrollbar(summary_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        # Download PDF Button
        ttk.Button(summary_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)

    def low_stock_report(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name, category, quantity, reorder_level FROM items WHERE section=?", (self.section,))
        items = c.fetchall()
        conn.close()

        report = f"{self.section.capitalize()} Low Stock Report:\n\n"
        low_items = []
        for name, category, qty, reorder in items:
            qty_num, qty_unit = self.app.parse_quantity(qty)
            reorder_num, reorder_unit = self.app.parse_quantity(reorder)
            qty_converted = self.app.convert_units(qty_num, qty_unit, reorder_unit)
            if qty_converted is not None and reorder_num is not None and qty_converted <= reorder_num:
                low_items.append((name, category, qty, reorder))

        if low_items:
            report += "Low Stock Items:\n"
            for item in low_items:
                report += f"Name: {item[0]}, Category: {item[1]}, Quantity: {item[2]}, Reorder Level: {item[3]}\n"
        else:
            report += "No items below reorder level.\n"

        # Create scrollable window with Treeview
        report_window = tk.Toplevel(self.app.root)
        report_window.title(f"{self.section.capitalize()} Low Stock Report")
        report_window.geometry("600x400")

        tree = ttk.Treeview(report_window, columns=("Name", "Category", "Quantity", "Reorder Level"), show="headings")
        tree.heading("Name", text="Name")
        tree.heading("Category", text="Category")
        tree.heading("Quantity", text="Quantity")
        tree.heading("Reorder Level", text="Reorder Level")
        tree.column("Name", width=150)
        tree.column("Category", width=100)
        tree.column("Quantity", width=100)
        tree.column("Reorder Level", width=100)

        # Insert data into Treeview
        if low_items:
            for item in low_items:
                tree.insert("", "end", values=item)
        else:
            tree.insert("", "end", values=("No items below reorder level", "", "", ""))

        # Scrollbar
        scrollbar = ttk.Scrollbar(report_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        # Download PDF Button
        ttk.Button(report_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)
    

    def category_analysis(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT category, COUNT(*), SUM(CAST(SUBSTR(quantity, 1, INSTR(quantity, ' ')-1) AS REAL)) FROM items WHERE section=? GROUP BY category", (self.section,))
        data = c.fetchall()
        conn.close()

        report = f"{self.section.capitalize()} Category Analysis:\n\n"
        if data:
            report += "Categories:\n"
            for cat, count, total_qty in data:
                report += f"Category: {cat}, Items: {count}, Total Quantity: {total_qty}\n"
        else:
            report += "No categories found.\n"

        # Create scrollable window with Treeview
        analysis_window = tk.Toplevel(self.app.root)
        analysis_window.title(f"{self.section.capitalize()} Category Analysis")
        analysis_window.geometry("600x400")

        tree = ttk.Treeview(analysis_window, columns=("Category", "Items", "Total Quantity"), show="headings")
        tree.heading("Category", text="Category")
        tree.heading("Items", text="Items")
        tree.heading("Total Quantity", text="Total Quantity")
        tree.column("Category", width=200)
        tree.column("Items", width=100)
        tree.column("Total Quantity", width=150)

        # Insert data into Treeview
        if data:
            for cat, count, total_qty in data:
                tree.insert("", "end", values=(cat, count, total_qty))
        else:
            tree.insert("", "end", values=("No categories found", "", ""))

        # Scrollbar
        scrollbar = ttk.Scrollbar(analysis_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        # Download PDF Button
        ttk.Button(analysis_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)

        # Optional: Keep the plots if desired (commented out for now)
        
        categories, counts, quantities = zip(*data) if data else ([], [], [])
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 4))
        ax1.pie(counts, labels=categories, autopct='%1.1f%%', startangle=90)
        ax1.set_title(f"{self.section.capitalize()} Items by Category")
        ax2.bar(categories, quantities, color="lightgreen")
        ax2.set_title(f"{self.section.capitalize()} Quantity by Category")
        ax2.set_xlabel("Category")
        ax2.set_ylabel("Total Quantity (numeric only)")
        plt.xticks(rotation=45, ha="right")
        canvas = FigureCanvasTkAgg(fig, master=analysis_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=1)
        

    def broken_items_report(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        # Remove section filter if not in schema; assume broken_items is lab-specific
        c.execute("SELECT student_id, student_name, item_name, report_date, status FROM broken_items")
        broken_items = c.fetchall()
        conn.close()

        report = f"{self.section.capitalize()} Broken Items Report:\n\n"
        if broken_items:
            report += "Broken Items:\n"
            for item in broken_items:
                report += f"Student ID: {item[0]}, Name: {item[1]}, Item: {item[2]}, Date: {item[3]}, Status: {item[4]}\n"
        else:
            report += "No broken items reported.\n"

        # Create scrollable window with Treeview
        report_window = tk.Toplevel(self.app.root)
        report_window.title(f"{self.section.capitalize()} Broken Items Report")
        report_window.geometry("600x400")

        tree = ttk.Treeview(report_window, columns=("Student ID", "Student Name", "Item", "Date", "Status"), show="headings")
        tree.heading("Student ID", text="Student ID")
        tree.heading("Student Name", text="Student Name")
        tree.heading("Item", text="Item")
        tree.heading("Date", text="Date")
        tree.heading("Status", text="Status")
        tree.column("Student ID", width=100)
        tree.column("Student Name", width=150)
        tree.column("Item", width=100)
        tree.column("Date", width=100)
        tree.column("Status", width=100)

        if broken_items:
            for item in broken_items:
                tree.insert("", "end", values=item)
        else:
            tree.insert("", "end", values=("", "", "No broken items reported", "", ""))

        scrollbar = ttk.Scrollbar(report_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Button(report_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)

    def predictive_reorder(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("""
            SELECT i.name, i.quantity, i.reorder_level, iss.issue_date 
            FROM items i 
            LEFT JOIN issuance iss ON i.id = iss.item_id 
            WHERE i.section=?
        """, (self.section,))
        data = c.fetchall()
        conn.close()

        report = f"{self.section.capitalize()} Predictive Reorder Suggestions (ARIMA):\n\n"
        items_dict = {}
        for name, qty, reorder, issue_date in data:
            if name not in items_dict:
                items_dict[name] = {"quantity": qty, "reorder": reorder or "0", "dates": []}
            if issue_date:
                try:
                    items_dict[name]["dates"].append(datetime.strptime(issue_date, "%Y-%m-%d %H:%M:%S"))
                except ValueError:
                    continue

        predictions = []
        for name, info in items_dict.items():
            qty_num, qty_unit = self.app.parse_quantity(info["quantity"])
            reorder_num, reorder_unit = self.app.parse_quantity(info["reorder"])
            dates = sorted(info["dates"])

            if qty_num is None or reorder_num is None:
                report += f"{name}: Invalid quantity or reorder data, skipping prediction.\n"
                predictions.append((name, info["quantity"], info["reorder"], "N/A"))
                continue

            if len(dates) >= 5:
                start_date = min(dates)
                end_date = max(dates, default=datetime.now())
                days = (end_date - start_date).days + 1
                usage = np.zeros(days)
                for date in dates:
                    day_idx = (date - start_date).days
                    usage[day_idx] += 1  # Assumes 1 unit; adjust if quantity_issued exists

                try:
                    model = sm.tsa.ARIMA(usage, order=(1, 1, 1))
                    results = model.fit()
                    forecast = results.forecast(steps=30)
                    avg_usage = np.nanmean(forecast)  # Handle NaN in forecast
                    if np.isnan(avg_usage) or avg_usage < 0:
                        avg_usage = len(dates) / days  # Fallback
                except Exception:
                    avg_usage = len(dates) / days  # Fallback to average over period
            else:
                avg_usage = len(dates) / 30 if dates else 0

            suggested_reorder = max(reorder_num, int(avg_usage * 1.5 * self.app.convert_units(1, qty_unit, reorder_unit)))
            report += f"{name}: Current {info['quantity']}, Current Reorder {info['reorder']}, Avg Usage/Day {avg_usage:.2f}, Suggested Reorder {suggested_reorder} {reorder_unit}\n"
            predictions.append((name, info["quantity"], info["reorder"], f"{suggested_reorder} {reorder_unit}"))

        if not predictions:
            report += "No predictions available.\n"

        report_window = tk.Toplevel(self.app.root)
        report_window.title(f"{self.section.capitalize()} Predictive Reorder Suggestions")
        report_window.geometry("600x400")

        tree = ttk.Treeview(report_window, columns=("Name", "Current Qty", "Reorder Level", "Suggested"), show="headings")
        tree.heading("Name", text="Name")
        tree.heading("Current Qty", text="Current Quantity")
        tree.heading("Reorder Level", text="Reorder Level")
        tree.heading("Suggested", text="Suggested Reorder")
        tree.column("Name", width=150)
        tree.column("Current Qty", width=100)
        tree.column("Reorder Level", width=100)
        tree.column("Suggested", width=150)

        for pred in predictions:
            tree.insert("", "end", values=pred)
        if not predictions:
            tree.insert("", "end", values=("No predictions available", "", "", ""))

        scrollbar = ttk.Scrollbar(report_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Button(report_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)
        
    def audit_report(self):
        # Fetch all data from the database
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT id, name, category, quantity, reorder_level, expiry_date, last_updated, section, purchase_unit FROM items WHERE section=?", (self.section,))
        items = c.fetchall()
        c.execute("SELECT i.name, p.price, p.last_updated FROM items i LEFT JOIN prices p ON i.id = p.item_id WHERE i.section=?", (self.section,))
        prices = c.fetchall()
        c.execute("SELECT inv.supplier, i.name, inv.quantity, inv.price, inv.invoice_date FROM invoices inv JOIN items i ON inv.item_id = i.id WHERE i.section=?", (self.section,))
        invoices = c.fetchall()
        c.execute("SELECT i.name, iss.person_name, iss.quantity_issued, iss.issue_date FROM issuance iss JOIN items i ON iss.item_id = i.id WHERE i.section=?", (self.section,))
        issuances = c.fetchall()
        c.execute("SELECT student_id, student_name, item_name, report_date, status FROM broken_items WHERE item_id IN (SELECT id FROM items WHERE section=?)", (self.section,))
        broken_items = c.fetchall()
        c.execute("SELECT action, i.name, details, timestamp FROM history h LEFT JOIN items i ON h.item_id = i.id WHERE i.section=?", (self.section,))
        history = c.fetchall()
        c.execute("SELECT id, username FROM users")
        users = c.fetchall()
        c.execute("SELECT subject, form, num_students, topic, subtopic, time, status FROM practical_reports")
        practical_reports = c.fetchall()
        c.execute("SELECT name, requirements FROM meal_templates")
        meal_templates = c.fetchall()
        c.execute("SELECT i.name, b.batch_number, b.quantity, b.unit_cost, b.received_date, b.expiry_date FROM batches b JOIN items i ON b.item_id = i.id WHERE i.section=?", (self.section,))
        batches = c.fetchall()
        conn.close()

        # Comprehensive report for PDF
        report_title = f"Comprehensive {self.section.capitalize()} Audit Report - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        report = f"{report_title}\n\n"

        # Items Section
        report += "Items:\n"
        if not items:
            report += "No items found.\n"
        for item in items:
            report += f"ID: {item[0]}, Name: {item[1]}, Category: {item[2]}, Quantity: {item[3]}, Reorder: {item[4]}, Expiry: {item[5]}, Updated: {item[6]}, Section: {item[7]}, Purchase Unit: {item[8]}\n"

        # Prices Section
        report += "\nPrices:\n"
        if not prices:
            report += "No price data available.\n"
        for price in prices:
            price_value = f"Ksh{float(price[1]):.2f}" if price[1] is not None else "KshN/A"
            report += f"Name: {price[0]}, Price: {price_value}, Updated: {price[2] or 'N/A'}\n"

        # Invoices Section
        report += "\nInvoices:\n"
        if not invoices:
            report += "No invoices found.\n"
        for inv in invoices:
            inv_price = f"Ksh{float(inv[3]):.2f}" if inv[3] is not None else "KshN/A"
            report += f"Supplier: {inv[0]}, Item: {inv[1]}, Quantity: {inv[2]}, Price: {inv_price}, Date: {inv[4] or 'N/A'}\n"

        # Issuance Section
        report += "\nIssuance:\n"
        if not issuances:
            report += "No issuance records found.\n"
        for iss in issuances:
            report += f"Item: {iss[0]}, Person: {iss[1]}, Quantity Issued: {iss[2]}, Date: {iss[3] or 'N/A'}\n"

        # Broken Items Section
        report += "\nBroken Items:\n"
        if not broken_items:
            report += "No broken items reported.\n"
        for broken in broken_items:
            report += f"Student ID: {broken[0]}, Student Name: {broken[1]}, Item: {broken[2]}, Report Date: {broken[3]}, Status: {broken[4]}\n"

        # History Section
        report += "\nHistory:\n"
        if not history:
            report += "No history records found.\n"
        for hist in history:
            report += f"Action: {hist[0]}, Item: {hist[1] or 'N/A'}, Details: {hist[2]}, Timestamp: {hist[3]}\n"

        # Users Section
        report += "\nUsers:\n"
        if not users:
            report += "No users found.\n"
        for user in users:
            report += f"ID: {user[0]}, Username: {user[1]}\n"

        # Practical Reports Section (lab only)
        if self.section == "lab":
            report += "\nPractical Reports:\n"
            if not practical_reports:
                report += "No practical reports found.\n"
            for prac in practical_reports:
                report += f"Subject: {prac[0]}, Form: {prac[1]}, Students: {prac[2]}, Topic: {prac[3]}, Subtopic: {prac[4]}, Time: {prac[5]}, Status: {prac[6]}\n"

        # Batches Section
        report += "\nBatches:\n"
        if not batches:
            report += "No batches found.\n"
        for batch in batches:
            report += f"Item: {batch[0]}, Batch Number: {batch[1]}, Quantity: {batch[2]}, Unit Cost: Ksh{float(batch[3]):.2f}, Received: {batch[4]}, Expiry: {batch[5]}\n"

        # Create Treeview window
        audit_window = tk.Toplevel(self.app.root)
        audit_window.title(f"Comprehensive {self.section.capitalize()} Audit Report")
        audit_window.geometry("1000x800")
        audit_window.columnconfigure(0, weight=1)
        audit_window.rowconfigure(0, weight=1)

        tree_frame = ttk.Frame(audit_window)
        tree_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=5)
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

        audit_tree = ttk.Treeview(tree_frame, columns=("Details",), show="tree headings")
        audit_tree.heading("Details", text="Audit Details")
        audit_tree.column("Details", stretch=tk.YES, width=900)
        audit_tree.grid(row=0, column=0, sticky="nsew")

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=audit_tree.yview)
        audit_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.grid(row=0, column=1, sticky="ns")

        # Populate Treeview
        def add_section(parent, title, data, format_func):
            section_node = audit_tree.insert(parent, "end", text=title, open=True)
            if not data:
                audit_tree.insert(section_node, "end", text="No data found.")
            else:
                for entry in data:
                    audit_tree.insert(section_node, "end", text=format_func(entry))

        format_item = lambda x: f"ID: {x[0]}, Name: {x[1]}, Category: {x[2]}, Qty: {x[3]}, Reorder: {x[4]}, Expiry: {x[5]}, Updated: {x[6]}, Section: {x[7]}, Purchase Unit: {x[8]}"
        format_price = lambda x: f"Name: {x[0]}, Price: {'Ksh' + f'{float(x[1]):.2f}' if x[1] is not None else 'KshN/A'}, Updated: {x[2] or 'N/A'}"
        format_invoice = lambda x: f"Supplier: {x[0]}, Item: {x[1]}, Qty: {x[2]}, Price: {'Ksh' + f'{float(x[3]):.2f}' if x[3] is not None else 'KshN/A'}, Date: {x[4] or 'N/A'}"
        format_issuance = lambda x: f"Item: {x[0]}, Person: {x[1]}, Qty Issued: {x[2]}, Date: {x[3] or 'N/A'}"
        format_broken = lambda x: f"Student ID: {x[0]}, Student Name: {x[1]}, Item: {x[2]}, Report Date: {x[3]}, Status: {x[4]}"
        format_history = lambda x: f"Action: {x[0]}, Item: {x[1] or 'N/A'}, Details: {x[2]}, Timestamp: {x[3]}"
        format_user = lambda x: f"ID: {x[0]}, Username: {x[1]}"
        format_practical = lambda x: f"Subject: {x[0]}, Form: {x[1]}, Students: {x[2]}, Topic: {x[3]}, Subtopic: {x[4]}, Time: {x[5]}, Status: {x[6]}"
        format_batch = lambda x: f"Item: {x[0]}, Batch: {x[1]}, Qty: {x[2]}, Cost: Ksh{float(x[3]):.2f}, Received: {x[4]}, Expiry: {x[5]}"

        add_section("", "Items", items, format_item)
        add_section("", "Prices", prices, format_price)
        add_section("", "Invoices", invoices, format_invoice)
        add_section("", "Issuance", issuances, format_issuance)
        add_section("", "Broken Items", broken_items, format_broken)
        add_section("", "History", history, format_history)
        add_section("", "Users", users, format_user)
        if self.section == "lab":
            add_section("", "Practical Reports", practical_reports, format_practical)
        add_section("", "Batches", batches, format_batch)

        # Download Button
        button_frame = ttk.Frame(audit_window)
        button_frame.grid(row=1, column=0, pady=10, sticky="ew")
        button_frame.columnconfigure(0, weight=1)
        ttk.Button(button_frame, text="Download PDF", command=lambda: self.download_audit_pdf(report)).grid(row=0, column=0, padx=5, pady=5)

        logging.info(f"Generated comprehensive {self.section} audit report")

    def download_audit_pdf(self, report):
        filename = f"{self.section}_audit_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Log the raw report for debugging
        logging.debug(f"Raw report content:\n{report}")

        # Split report into lines
        lines = report.split("\n")
        if not lines:
            logging.error("Report is empty, no lines to process.")
            elements.append(Paragraph("No data available.", styles['Normal']))
        else:
            # Add title
            elements.append(Paragraph(lines[0], styles['Title']))
            elements.append(Paragraph("<br/>", styles['Normal']))
            logging.debug(f"Added title: {lines[0]}")

            current_section = None
            table_data = []
            headers = None

            for line in lines[1:]:
                line = line.strip()
                if not line:
                    continue  # Skip empty lines

                # Check if this is a section header
                if line.endswith(":") and "No " not in line:
                    # Finish previous section
                    if current_section and table_data:
                        logging.debug(f"Building table for {current_section} with {len(table_data)} rows")
                        table = Table([headers] + table_data)
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ]))
                        elements.append(table)
                        elements.append(Paragraph("<br/>", styles['Normal']))
                        table_data = []
                        headers = None
                    elif current_section:
                        logging.debug(f"Adding plain text for {current_section}")
                        elements.append(Paragraph(current_section, styles['Heading2']))

                    # Start new section
                    current_section = line
                    elements.append(Paragraph(line, styles['Heading2']))
                    logging.debug(f"New section: {line}")

                # Handle "No data" messages or data rows
                elif current_section:
                    if "No " in line:
                        elements.append(Paragraph(line, styles['Normal']))
                        logging.debug(f"Added no-data line: {line}")
                    else:
                        # Parse data rows
                        if current_section == "Items:":
                            headers = ["ID", "Name", "Category", "Quantity", "Reorder", "Expiry", "Updated", "Section", "Purchase Unit"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "Prices:":
                            headers = ["Name", "Price", "Updated"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "Invoices:":
                            headers = ["Supplier", "Item", "Quantity", "Price", "Date"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "Issuance:":
                            headers = ["Item", "Person", "Quantity Issued", "Date"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "Broken Items:":
                            headers = ["Student ID", "Student Name", "Item", "Report Date", "Status"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "History:":
                            headers = ["Action", "Item", "Details", "Timestamp"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "Users:":
                            headers = ["ID", "Username"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "Practical Reports:":
                            headers = ["Subject", "Form", "Students", "Topic", "Subtopic", "Time", "Status"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "Meal Templates:":
                            headers = ["Name", "Requirements"]
                            parts = line.split(", ", 1)  # Only split on first comma for requirements
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "Batches:":
                            headers = ["Item", "Batch Number", "Quantity", "Unit Cost", "Received", "Expiry"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        logging.debug(f"Added row to {current_section}: {line}")

            # Handle the last section
            if current_section and table_data:
                logging.debug(f"Building final table for {current_section} with {len(table_data)} rows")
                table = Table([headers] + table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)

        # Build the PDF
        try:
            doc.build(elements)
            logging.info(f"PDF built successfully: {filename}")
        except Exception as e:
            logging.error(f"Failed to build PDF: {str(e)}")
            messagebox.showerror("Error", f"Failed to generate PDF: {str(e)}")
            return

        messagebox.showinfo("Success", f"Audit report saved as {filename}")
        webbrowser.open(filename)
        logging.info(f"Downloaded {self.section} audit report as {filename}")

    
    def manage_practical_reports(self):
        report_window = tk.Toplevel(self.app.root)
        report_window.title("Practical Reports")
        report_window.geometry("900x700")  # Adjusted initial size
        report_window.columnconfigure(1, weight=1)
        report_window.columnconfigure(2, weight=1)
        report_window.rowconfigure(8, weight=1)

        # --- Input Fields Frame ---
        input_frame = ttk.LabelFrame(report_window, text="Report Details")
        input_frame.grid(row=0, column=0, columnspan=3, padx=10, pady=10, sticky="ew")
        input_frame.columnconfigure(1, weight=1)

        ttk.Label(input_frame, text="Subject:", width=20, anchor="w").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.report_subject = ttk.Combobox(input_frame, values=["Physics", "Chemistry", "Biology"])
        self.report_subject.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(input_frame, text="Form (e.g., Form 1):", width=20, anchor="w").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.report_form = ttk.Entry(input_frame)
        self.report_form.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(input_frame, text="Number of Students:", width=20, anchor="w").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.report_students = ttk.Entry(input_frame)
        self.report_students.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(input_frame, text="Topic:", width=20, anchor="w").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.report_topic = ttk.Entry(input_frame)
        self.report_topic.grid(row=3, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(input_frame, text="Subtopic:", width=20, anchor="w").grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.report_subtopic = ttk.Entry(input_frame)
        self.report_subtopic.grid(row=4, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(input_frame, text="Time (YYYY-MM-DD HH:MM):", width=20, anchor="w").grid(row=5, column=0, padx=5, pady=5, sticky="w")
        self.report_time = ttk.Entry(input_frame)
        self.report_time.grid(row=5, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(input_frame, text="Status:", width=20, anchor="w").grid(row=6, column=0, padx=5, pady=5, sticky="w")
        self.report_status = ttk.Combobox(input_frame, values=["Done", "Pending"])
        self.report_status.grid(row=6, column=1, padx=5, pady=5, sticky="ew")
        self.report_status.set("Done")

        # --- Buttons Frame ---
        button_frame = ttk.Frame(report_window)
        button_frame.grid(row=7, column=0, columnspan=3, pady=10, sticky="ew")
        button_frame.columnconfigure(0, weight=1)
        button_frame.columnconfigure(1, weight=1)
        button_frame.columnconfigure(2, weight=1)

        ttk.Button(button_frame, text="Add Report", command=self.add_practical_report).grid(row=0, column=0, padx=5, sticky="ew")
        ttk.Button(button_frame, text="Download Report", command=self.download_practical_report).grid(row=0, column=1, padx=5, sticky="ew")
        ttk.Button(button_frame, text="Show Timeline", command=self.show_practical_timeline).grid(row=0, column=2, padx=5, sticky="ew")

        # --- Report Treeview ---
        self.report_tree = ttk.Treeview(report_window, columns=("ID", "Subject", "Form", "Students", "Topic", "Subtopic", "Time", "Status"), show="headings")
        for col in self.report_tree["columns"]:
            self.report_tree.heading(col, text=col)
            self.report_tree.column(col, width=100, stretch=tk.YES) # Set initial width and allow stretching

        self.report_tree.grid(row=8, column=0, columnspan=3, padx=10, pady=10, sticky="nsew")

        self.load_practical_reports()

    def show_practical_timeline(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT subject, time FROM practical_reports")
        data = c.fetchall()
        conn.close()
    
        if not data:
            messagebox.showwarning("Warning", "No practical reports to display.")
            return
    
        subjects, times = zip(*data)
        start_times = [datetime.strptime(t, "%Y-%m-%d %H:%M") for t in times]
        fig, ax = plt.subplots(figsize=(10, 4))
        ax.barh(subjects, [2] * len(subjects), left=start_times, height=0.5, color="skyblue")  # Assume 2-hour duration
        ax.set_title("Practical Schedule Timeline")
        ax.set_xlabel("Date and Time")
        ax.set_ylabel("Subject")
        plt.tight_layout()
    
        window = tk.Toplevel(self.app.root)
        window.title("Practical Timeline")
        canvas = FigureCanvasTkAgg(fig, master=window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=1)


    def download_practical_report(self):
        # Ask user for the export format
        format_choice = tk.simpledialog.askstring("Download Format", "Enter format (csv, excel, pdf):", initialvalue="csv")
        if not format_choice:
            return
    
        format_choice = format_choice.lower()
        if format_choice not in ["csv", "excel", "pdf"]:
            messagebox.showerror("Error", "Invalid format. Choose 'csv', 'excel', or 'pdf'.")
            return

        # Fetch data from practical_reports
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT id, subject, form, num_students, topic, subtopic, time, status FROM practical_reports")
        rows = c.fetchall()
        conn.close()

        if not rows:
            messagebox.showwarning("Warning", "No practical reports to download.")
            return

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"practical_reports_{timestamp}.{format_choice}"

        # Export based on format
        if format_choice == "csv":
            with open(filename, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["ID", "Subject", "Form", "Number of Students", "Topic", "Subtopic", "Time", "Status"])
                writer.writerows(rows)
            messagebox.showinfo("Success", f"Practical reports exported to {filename}")

        elif format_choice == "excel":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Practical Reports"
        
            headers = ["ID", "Subject", "Form", "Number of Students", "Topic", "Subtopic", "Time", "Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
        
            for row_idx, row in enumerate(rows, 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
            wb.save(filename)
            messagebox.showinfo("Success", f"Practical reports exported to {filename}")

        elif format_choice == "pdf":
            doc = SimpleDocTemplate(filename, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
        
            elements.append(Paragraph("Laboratory Practical Reports", styles['Title']))
            elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            elements.append(Paragraph("<br/><br/>", styles['Normal']))
        
            data = [["ID", "Subject", "Form", "Students", "Topic", "Subtopic", "Time", "Status"]] + list(rows)
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), '#808080'),
                ('TEXTCOLOR', (0, 0), (-1, 0), '#FFFFFF'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), '#F5F5DC'),
                ('GRID', (0, 0), (-1, -1), 1, '#000000'),
            ]))
            elements.append(table)
        
            doc.build(elements)
            messagebox.showinfo("Success", f"Practical reports exported to {filename}")
            webbrowser.open(filename)  # Optional: Opens the PDF after creation

        logging.info(f"Downloaded practical reports as {filename}")

    def add_practical_report(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("CREATE TABLE IF NOT EXISTS practical_reports (id INTEGER PRIMARY KEY, subject TEXT, form TEXT, num_students INTEGER, topic TEXT, subtopic TEXT, time TEXT, status TEXT)")
        c.execute("INSERT INTO practical_reports (subject, form, num_students, topic, subtopic, time, status) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (self.report_subject.get(), self.report_form.get(), int(self.report_students.get()), self.report_topic.get(), self.report_subtopic.get(), self.report_time.get(), self.report_status.get()))
        conn.commit()
        conn.close()
        self.load_practical_reports()
        logging.info(f"Added practical report for {self.report_subject.get()}")

    def load_practical_reports(self):
        self.report_tree.delete(*self.report_tree.get_children())
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT id, subject, form, num_students, topic, subtopic, time, status FROM practical_reports")
        for row in c.fetchall():
            self.report_tree.insert("", "end", values=row)
        conn.close()


    def gmp_analysis(self):
        """
        Good Manufacturing Practices (GMP) Analysis for Lab:
        - Checks expiry dates for safety (chemicals/equipment).
        - Ensures sufficient stock and reorder compliance.
        - Verifies batch traceability for chemicals.
        - Flags items for disposal if expired.
        """
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("""
            SELECT i.name, i.quantity, i.reorder_level, i.expiry_date, i.purchase_unit, 
                   COUNT(b.id) as batch_count, p.price 
            FROM items i 
            LEFT JOIN batches b ON i.id = b.item_id 
            LEFT JOIN prices p ON i.id = p.item_id 
            WHERE i.section=? 
            GROUP BY i.id, i.name, i.quantity, i.reorder_level, i.expiry_date, i.purchase_unit, p.price
        """, (self.section,))
        items = c.fetchall()
        conn.close()

        now = datetime.now()
        report = f"{self.section.capitalize()} GMP Analysis:\n\n"
        expiry_issues = []
        stock_issues = []
        traceability_issues = []
        disposal_items = []

        for name, qty, reorder, expiry_date, purchase_unit, batch_count, price in items:
            qty_num, qty_unit = self.app.parse_quantity(qty)
            reorder_num, _ = self.app.parse_quantity(reorder or "0")  # Default to "0" if None
            pu_num, _ = self.app.parse_quantity(purchase_unit) if purchase_unit else (1, qty_unit)
            num_units = qty_num / pu_num if pu_num and qty_num is not None else 0
            reorder_units = reorder_num / pu_num if pu_num and reorder_num is not None else 0

            # 1. Expiry Check
            if expiry_date:
                try:
                    expiry = datetime.strptime(expiry_date, "%Y-%m-%d")
                    days_left = (expiry - now).days
                    if days_left <= 14:  # Lab-specific threshold
                        expiry_issues.append((name, f"Expires in {days_left} days", expiry_date))
                        if days_left <= 0:
                            disposal_items.append((name, expiry_date, f"{num_units:.2f} x {purchase_unit}"))
                except ValueError:
                    expiry_issues.append((name, "Invalid expiry date", expiry_date))

            # 2. Stock Compliance
            if qty_num is not None and reorder_num is not None and num_units <= reorder_units:
                stock_issues.append((name, f"{num_units:.2f} x {purchase_unit}", f"{reorder_units:.2f} x {purchase_unit}"))

            # 3. Traceability (Lab-specific: chemicals need batches)
            if batch_count == 0 and purchase_unit and ("ml" in purchase_unit.lower() or "g" in purchase_unit.lower()):
                traceability_issues.append((name, purchase_unit))

        # Compile Report Sections
        report += "1. Expiry and Safety Check:\n"
        report += "\n".join(f"{item[0]}: {item[1]} ({item[2]})" for item in expiry_issues) + "\n" if expiry_issues else "No expiry issues detected.\n"

        report += "\n2. Stock Compliance:\n"
        report += "\n".join(f"{item[0]}: Qty {item[1]} <= Reorder {item[2]}" for item in stock_issues) + "\n" if stock_issues else "All items above reorder levels.\n"

        report += "\n3. Traceability (Batch Tracking):\n"
        report += "\n".join(f"{item[0]}: No batch records (Chemical: {item[1]})" for item in traceability_issues) + "\n" if traceability_issues else "Batch tracking adequate.\n"

        report += "\n4. Items for Disposal:\n"
        report += "\n".join(f"{item[0]}: Expired on {item[1]}, Qty: {item[2]}" for item in disposal_items) + "\n" if disposal_items else "No items require disposal.\n"

        # Recommendations
        recommendations = []
        if expiry_issues:
            recommendations.append("Replace or dispose of items nearing expiry (chemical safety critical).")
        if stock_issues:
            recommendations.append("Order additional stock for lab items below reorder levels.")
        if traceability_issues:
            recommendations.append("Implement batch tracking for chemicals.")
        if disposal_items:
            recommendations.append("Dispose of expired items per lab safety protocols.")
        report += "\nRecommendations:\n"
        report += "\n".join(recommendations) + "\n" if recommendations else "Maintain current practices.\n"

        # Create scrollable window with Treeview
        report_window = tk.Toplevel(self.app.root)
        report_window.title(f"{self.section.capitalize()} GMP Analysis")
        report_window.geometry("800x500")

        # Main Treeview for overview
        tree = ttk.Treeview(report_window, columns=("Category", "Item", "Details"), show="headings")
        tree.heading("Category", text="Category")
        tree.heading("Item", text="Item")
        tree.heading("Details", text="Details")
        tree.column("Category", width=200)
        tree.column("Item", width=200)
        tree.column("Details", width=350)

        # Populate Treeview
        if expiry_issues:
            tree.insert("", "end", values=("Expiry Issues", "", ""))
            for item in expiry_issues:
                tree.insert("", "end", values=("", item[0], f"{item[1]} ({item[2]})"))
        else:
            tree.insert("", "end", values=("Expiry Issues", "No issues", ""))

        if stock_issues:
            tree.insert("", "end", values=("Stock Compliance", "", ""))
            for item in stock_issues:
                tree.insert("", "end", values=("", item[0], f"Qty {item[1]} <= Reorder {item[2]}"))
        else:
            tree.insert("", "end", values=("Stock Compliance", "All adequate", ""))

        if traceability_issues:
            tree.insert("", "end", values=("Traceability", "", ""))
            for item in traceability_issues:
                tree.insert("", "end", values=("", item[0], f"No batch records (Chemical: {item[1]})"))
        else:
            tree.insert("", "end", values=("Traceability", "Adequate", ""))

        if disposal_items:
            tree.insert("", "end", values=("Disposal Items", "", ""))
            for item in disposal_items:
                tree.insert("", "end", values=("", item[0], f"Expired on {item[1]}, Qty: {item[2]}"))
        else:
            tree.insert("", "end", values=("Disposal Items", "None required", ""))

        if recommendations:
            tree.insert("", "end", values=("Recommendations", "", ""))
            for rec in recommendations:
                tree.insert("", "end", values=("", "", rec))
        else:
            tree.insert("", "end", values=("Recommendations", "Maintain current practices", ""))

        # Scrollbar
        scrollbar = ttk.Scrollbar(report_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        # Download PDF Button
        ttk.Button(report_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)
        logging.info("Performed lab GMP analysis")