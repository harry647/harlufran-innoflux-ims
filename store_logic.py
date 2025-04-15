# store_logic.py

import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox
import csv
from plyer import notification
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl
from openpyxl.styles import Font
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate
from reportlab.lib import colors
from datetime import datetime, timedelta
import webbrowser
import statsmodels.api as sm
import numpy as np
import logging
import os


# Replace the existing logging.basicConfig call
log_dir = os.path.join(os.getenv("APPDATA"), "InventoryManagementSystem", "logs")
os.makedirs(log_dir, exist_ok=True)
log_file = os.path.join(log_dir, "inventory_app.log")

logging.basicConfig(
    filename=log_file,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class StoreLogic:
    def __init__(self, app, store_tab):
        self.app = app
        self.store_tab = store_tab
        self.section = "store"
        self.store_tree = None  # Initialize here
        self.price_window = None
        self.price_tree = None
        self.price_entry = None
        self.invoice_window = None
        self.supplier_entry = None
        self.invoice_category = None
        self.invoice_items_tree = None
        self.invoice_quantity = None
        self.invoice_tree = None
        self.total_label = None
        self.invoice_items = []
        self.store_name = None
        self.store_category = None
        self.store_quantity = None
        self.store_reorder = None
        self.store_expiry = None
        self.issue_person = None
        self.issue_item = None
        self.issue_quantity = None
        self.store_search = None

        self.setup_store_tab()
        self.load_items(self.store_tree)
        self.check_reorder()
        self.check_expiry()

    
    def setup_store_tab(self):
        # Main Container Frame for Input and Issue Actions
        top_frame = ttk.Frame(self.store_tab)
        top_frame.pack(fill="x", padx=10, pady=5)
        top_frame.columnconfigure(0, weight=1)  # Input Frame
        top_frame.columnconfigure(1, weight=1)  # Issue Frame
        top_frame.columnconfigure(2, weight=1)  # Action Frame

        # Add/Edit Item Frame (Left Side)
        input_frame = ttk.LabelFrame(top_frame, text="Add/Edit Item")
        input_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        input_frame.columnconfigure(0, weight=0)
        input_frame.columnconfigure(1, weight=1)

        ttk.Label(input_frame, text="Name:").grid(row=0, column=0, padx=5, pady=2, sticky="w")
        self.store_name = ttk.Entry(input_frame)
        self.store_name.grid(row=0, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(input_frame, text="Category:").grid(row=1, column=0, padx=5, pady=2, sticky="w")
        self.store_category = ttk.Combobox(input_frame, values=self.get_categories())
        self.store_category.grid(row=1, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(input_frame, text="Purchase Unit (e.g., '50 kg'):").grid(row=2, column=0, padx=5, pady=2, sticky="w")
        self.store_purchase_unit = ttk.Entry(input_frame)
        self.store_purchase_unit.grid(row=2, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(input_frame, text="Quantity (number of purchase units, e.g., '23'):").grid(row=3, column=0, padx=5, pady=2, sticky="w")
        self.store_quantity = ttk.Entry(input_frame)
        self.store_quantity.grid(row=3, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(input_frame, text="Reorder Level (number of purchase units, e.g., '2'):").grid(row=4, column=0, padx=5, pady=2, sticky="w")
        self.store_reorder = ttk.Entry(input_frame)
        self.store_reorder.grid(row=4, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(input_frame, text="Expiry Date (YYYY-MM-DD):").grid(row=5, column=0, padx=5, pady=2, sticky="w")
        self.store_expiry = ttk.Entry(input_frame)
        self.store_expiry.grid(row=5, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(input_frame, text="Price per Purchase Unit (Ksh):").grid(row=6, column=0, padx=5, pady=2, sticky="w")
        self.store_price = ttk.Entry(input_frame)
        self.store_price.grid(row=6, column=1, padx=5, pady=2, sticky="ew")

        ttk.Button(input_frame, text="Add Item", command=self.add_store_item).grid(row=6, column=0, padx=5, pady=5, sticky="ew")
        ttk.Button(input_frame, text="Update Items", command=self.update_store_items).grid(row=6, column=1, padx=5, pady=5, sticky="ew")
            # Issue Item Frame (Right Side)
        issue_frame = ttk.LabelFrame(top_frame, text="Issue Item")
        issue_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
        issue_frame.columnconfigure(0, weight=0)
        issue_frame.columnconfigure(1, weight=1)

        ttk.Label(issue_frame, text="Person Name:").grid(row=0, column=0, padx=5, pady=2, sticky="w")
        self.issue_person = ttk.Entry(issue_frame)
        self.issue_person.grid(row=0, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(issue_frame, text="Item Name:").grid(row=1, column=0, padx=5, pady=2, sticky="w")
        self.issue_item = ttk.Combobox(issue_frame, values=self.get_item_names())
        self.issue_item.grid(row=1, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(issue_frame, text="Quantity (e.g., '3 units'):").grid(row=2, column=0, padx=5, pady=2, sticky="w")
        self.issue_quantity = ttk.Entry(issue_frame)
        self.issue_quantity.grid(row=2, column=1, padx=5, pady=2, sticky="ew")

        ttk.Button(issue_frame, text="Issue Item", command=self.issue_store_item).grid(row=3, column=1, padx=5, pady=5, sticky="ew")

        # Search Frame
        search_frame = ttk.Frame(self.store_tab)
        search_frame.pack(fill="x", padx=10, pady=5)
        search_frame.columnconfigure(2, weight=1)

        # Filter selection
        ttk.Label(search_frame, text="Search by:").pack(side="left", padx=(0, 5))
        self.search_filter = ttk.Combobox(search_frame, 
                                        values=["Name", "Category", "ID", "Quantity"], 
                                        state="readonly")
        self.search_filter.set("Name")  # Default filter
        self.search_filter.pack(side="left", padx=5)

        # Search entry
        ttk.Label(search_frame, text="Search:").pack(side="left")
        self.store_search = ttk.Entry(search_frame)
        self.store_search.pack(side="left", padx=5, fill="x", expand=True)
        self.store_search.bind("<KeyRelease>", lambda e: self.search_items())

        # Treeview for Inventory Display (removed Price column)
        self.store_tree = ttk.Treeview(self.store_tab, 
                                     columns=("ID", "Name", "Category", "Quantity", 
                                            "Reorder", "Expiry", "Last Updated"), 
                                     show="headings")
        self.store_tree.heading("ID", text="ID")
        self.store_tree.heading("Name", text="Name")
        self.store_tree.heading("Category", text="Category")
        self.store_tree.heading("Quantity", text="Quantity")
        self.store_tree.heading("Reorder", text="Reorder Level")
        self.store_tree.heading("Expiry", text="Expiry Date")
        self.store_tree.heading("Last Updated", text="Last Updated")
        self.store_tree.pack(fill="both", expand=1, padx=10, pady=5)
        self.store_tree.bind("<Button-1>", self.select_item)
        self.store_tree.bind("<Enter>", lambda e: self.app.show_tooltip(e, "Click to edit item"))
        self.store_tree.bind("<Leave>", self.app.hide_tooltip)
        
        for col in ("ID", "Name", "Category", "Quantity", "Reorder", "Expiry", "Last Updated"):
            self.store_tree.column(col, stretch=tk.YES)

        action_frame = ttk.LabelFrame(top_frame, text="More Actions")
        action_frame.grid(row=0, column=2, padx=5, pady=5, sticky="nsew")

        # Row 0: Core Actions + Supplier/Feeding
        for i in range(5):
            action_frame.columnconfigure(i, weight=1)
        ttk.Button(action_frame, text="Delete Item", command=self.delete_store_item).grid(row=0, column=0, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Update Prices", command=self.update_prices_window).grid(row=0, column=1, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Create Invoice", command=self.create_invoice_window).grid(row=0, column=2, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Manage Daily Suppliers", command=self.manage_daily_suppliers).grid(row=0, column=3, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Student Feeding Program", command=self.student_feeding_program).grid(row=0, column=4, padx=3, pady=5, sticky="ew")

        # Row 1: Export and Reporting Actions
        ttk.Button(action_frame, text="Export to CSV", command=self.export_to_csv).grid(row=1, column=0, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Export to Excel", command=self.export_to_excel).grid(row=1, column=1, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Export to PDF", command=self.export_to_pdf).grid(row=1, column=2, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Show Graph", command=self.show_graph).grid(row=1, column=3, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Audit Report", command=self.audit_report).grid(row=1, column=4, padx=3, pady=5, sticky="ew")

        # Row 2: Advanced Stock Management Features (updated)
        ttk.Button(action_frame, text="Stock Valuation", command=self.calculate_stock_value).grid(row=2, column=0, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="ABC Analysis", command=self.abc_analysis).grid(row=2, column=1, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Safety Stock", command=self.calculate_safety_stock).grid(row=2, column=2, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="EOQ", command=self.calculate_eoq).grid(row=2, column=3, padx=3, pady=5, sticky="ew")
        ttk.Button(action_frame, text="Batch Tracking", command=self.manage_batches).grid(row=2, column=4, padx=3, pady=5, sticky="ew")
        
        # Row 3: Add GMP Analysis
        ttk.Button(action_frame, text="GMP Analysis", command=self.gmp_analysis).grid(row=3, column=0, padx=3, pady=5, sticky="ew")
     

    def update_prices_window(self):
        self.price_window = tk.Toplevel(self.app.root)
        self.price_window.title("Update Item Prices")
        self.price_window.geometry("600x400")
        self.price_window.columnconfigure(0, weight=1)
        self.price_window.rowconfigure(1, weight=1)

        self.price_tree = ttk.Treeview(self.price_window, columns=("ID", "Name", "Purchase Unit", "Price", "Last Updated"), show="headings")
        self.price_tree.heading("ID", text="ID")
        self.price_tree.heading("Name", text="Name")
        self.price_tree.heading("Purchase Unit", text="Purchase Unit")
        self.price_tree.heading("Price", text="Price per Purchase Unit (Ksh)")
        self.price_tree.heading("Last Updated", text="Last Updated")
        for col in ("ID", "Name", "Purchase Unit", "Price", "Last Updated"):
            self.price_tree.column(col, stretch=tk.YES)

        scrollbar = ttk.Scrollbar(self.price_window, orient="vertical", command=self.price_tree.yview)
        self.price_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.price_tree.pack(fill="both", expand=1, padx=10, pady=5)

        ttk.Label(self.price_window, text="Price per Purchase Unit (Ksh):").pack(pady=5)
        self.price_entry = ttk.Entry(self.price_window)
        self.price_entry.pack(pady=5)
        self.price_entry.bind("<Return>", self.update_price)  # Bind Enter key

        ttk.Button(self.price_window, text="Update Price", command=self.update_price).pack(pady=10)

        self.price_tree.bind("<Button-1>", self.select_price_item)
        self.load_prices()

    

    def load_prices(self):
        self.price_tree.delete(*self.price_tree.get_children())
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("""
            SELECT i.id, i.name, i.purchase_unit, p.price, p.last_updated 
            FROM items i 
            LEFT JOIN prices p ON i.id = p.item_id 
            WHERE i.section=?
        """, (self.section,))
        for row in c.fetchall():
            self.price_tree.insert("", "end", values=(row[0], row[1], row[2] if row[2] else "N/A", row[3] if row[3] is not None else "", row[4]))
        conn.close()

    def select_price_item(self, event):
        selected = self.price_tree.selection()
        if selected:
            values = self.price_tree.item(selected)["values"]
            self.price_entry.delete(0, tk.END)
            self.price_entry.insert(0, values[3] if values[3] else "")  # Price per purchase unit

    def update_price(self):
        selected = self.price_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Select an item to update its price")
            return
        
        item_id = self.price_tree.item(selected)["values"][0]
        price = self.price_entry.get().strip()  # Price per purchase unit
        
        if not price:
            messagebox.showerror("Error", "Price cannot be empty")
            return
        
        try:
            price = float(price)
            if price < 0:
                messagebox.showerror("Error", "Price cannot be negative")
                return
            
            last_updated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn = sqlite3.connect(self.app.get_db_path())
            c = conn.cursor()
            
            # Get purchase unit for logging
            c.execute("SELECT name, purchase_unit FROM items WHERE id=?", (item_id,))
            name, purchase_unit = c.fetchone()
            
            # Update price (interpreted as per purchase unit)
            c.execute("""
                INSERT OR REPLACE INTO prices (item_id, price, last_updated) 
                VALUES (?, ?, ?)
            """, (item_id, price, last_updated))
            
            c.execute("""
                INSERT INTO history (action, item_id, details, timestamp) 
                VALUES (?, ?, ?, ?)
            """, ("Update Price", item_id, f"Price set to Ksh{price:.2f} per {purchase_unit}", last_updated))
            
            conn.commit()
            conn.close()
            
            self.load_prices()
            messagebox.showinfo("Success", "Price updated successfully")
            logging.info(f"Updated price for item ID {item_id} ({name}) to Ksh{price:.2f} per {purchase_unit}")
        
        except ValueError:
            messagebox.showerror("Error", "Price must be a valid number (e.g., 12.50)")
        except sqlite3.Error as e:
            messagebox.showerror("Error", f"Database error: {str(e)}")
            logging.error(f"Database error updating price for item ID {item_id}: {str(e)}")
        finally:
            if 'conn' in locals():
                conn.close()

    def update_purchase_unit(self, item_id, new_purchase_unit):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT quantity, purchase_unit FROM items WHERE id=?", (item_id,))
        current_qty, old_purchase_unit = c.fetchone()
        curr_num, curr_unit = self.app.parse_quantity(current_qty)
        old_pu_num, _ = self.app.parse_quantity(old_purchase_unit)
        new_pu_num, new_pu_unit = self.app.parse_quantity(new_purchase_unit)
        num_units = curr_num / old_pu_num  # Current number of purchase units
        new_total_qty = num_units * new_pu_num
        c.execute("UPDATE items SET quantity=?, purchase_unit=? WHERE id=?", (f"{new_total_qty} {new_pu_unit}", new_purchase_unit, item_id))
        conn.commit()
        conn.close()

    def create_invoice_window(self):
        self.invoice_window = tk.Toplevel(self.app.root)
        self.invoice_window.title("Create Supplier Invoice")
        self.invoice_window.geometry("800x600")
        self.invoice_window.columnconfigure(0, weight=1)
        self.invoice_window.columnconfigure(1, weight=1)
        self.invoice_window.rowconfigure(2, weight=1)
        self.invoice_window.rowconfigure(5, weight=1)

        ttk.Label(self.invoice_window, text="Supplier Name:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.supplier_entry = ttk.Entry(self.invoice_window)
        self.supplier_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(self.invoice_window, text="Category Filter:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.invoice_category = ttk.Combobox(self.invoice_window, values=["All"] + self.get_categories())
        self.invoice_category.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        self.invoice_category.set("All")
        self.invoice_category.bind("<<ComboboxSelected>>", self.load_invoice_items)

        self.invoice_items_tree = ttk.Treeview(self.invoice_window, columns=("ID", "Name", "Quantity", "Price"), show="headings")
        self.invoice_items_tree.heading("ID", text="ID")
        self.invoice_items_tree.heading("Name", text="Name")
        self.invoice_items_tree.heading("Quantity", text="Quantity")
        self.invoice_items_tree.heading("Price", text="Price (Ksh)")
        for col in ("ID", "Name", "Quantity", "Price"):
            self.invoice_items_tree.column(col, stretch=tk.YES)

        scrollbar = ttk.Scrollbar(self.invoice_window, orient="vertical", command=self.invoice_items_tree.yview)
        self.invoice_items_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.grid(row=2, column=2, sticky="ns", pady=5)
        self.invoice_items_tree.grid(row=2, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")

        ttk.Label(self.invoice_window, text="Quantity (e.g., '5 units'):").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.invoice_quantity = ttk.Entry(self.invoice_window)
        self.invoice_quantity.grid(row=3, column=1, padx=5, pady=5, sticky="ew")
        self.invoice_quantity.bind("<Return>", self.add_to_invoice) # Bind Enter key

        ttk.Button(self.invoice_window, text="Add to Invoice", command=self.add_to_invoice).grid(row=4, column=1, pady=10, sticky="ew")

        self.invoice_tree = ttk.Treeview(self.invoice_window, columns=("ID", "Name", "Quantity", "Price", "Total"), show="headings")
        self.invoice_tree.heading("ID", text="ID")
        self.invoice_tree.heading("Name", text="Name")
        self.invoice_tree.heading("Quantity", text="Quantity")
        self.invoice_tree.heading("Price", text="Price (Ksh)")
        self.invoice_tree.heading("Total", text="Total (Ksh)")
        for col in ("ID", "Name", "Quantity", "Price", "Total"):
            self.invoice_tree.column(col, stretch=tk.YES)
            self.invoice_tree.grid(row=5, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")
        
        ttk.Label(self.invoice_window, text="Total Amount:").grid(row=6, column=0, padx=5, pady=5, sticky="w")
        self.total_label = ttk.Label(self.invoice_window, text="Ksh 0.00")
        self.total_label.grid(row=6, column=1, padx=5, pady=5, sticky="ew")

        ttk.Button(self.invoice_window, text="Save & Print Invoice", command=self.save_and_print_invoice).grid(row=7, column=1, pady=10, sticky="ew")

        self.invoice_items = []
        self.load_invoice_items(None)

    def load_invoice_items(self, event):
        self.invoice_items_tree.delete(*self.invoice_items_tree.get_children())
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        category = self.invoice_category.get()
        if category == "All":
            c.execute("SELECT i.id, i.name, i.purchase_unit, p.price FROM items i LEFT JOIN prices p ON i.id = p.item_id WHERE i.section=?", (self.section,))
        else:
            c.execute("SELECT i.id, i.name, i.purchase_unit, p.price FROM items i LEFT JOIN prices p ON i.id = p.item_id WHERE i.section=? AND i.category=?", (self.section, category))
        for row in c.fetchall():
            self.invoice_items_tree.insert("", "end", values=(row[0], row[1], row[2] if row[2] else "N/A", row[3] if row[3] else 0.0))
        conn.close()

    def add_to_invoice(self):
        selected = self.invoice_items_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Select an item to add to the invoice")
            return
        
        item_id = self.invoice_items_tree.item(selected)["values"][0]
        name = self.invoice_items_tree.item(selected)["values"][1]
        purchase_unit = self.invoice_items_tree.item(selected)["values"][2]
        price_per_unit = self.invoice_items_tree.item(selected)["values"][3]  # Price per purchase unit
        quantity = self.invoice_quantity.get()  # Number of purchase units
        
        if not quantity:
            messagebox.showerror("Error", "Enter a quantity")
            return
        
        qty_num = float(quantity)
        total = price_per_unit * qty_num  # Total cost = price per purchase unit * number of units
        self.invoice_items.append((item_id, name, f"{quantity} x {purchase_unit}", price_per_unit, total))
        self.invoice_tree.insert("", "end", values=(item_id, name, f"{quantity} x {purchase_unit}", price_per_unit, total))
        self.update_invoice_total()


    def update_invoice_total(self):
        total = sum(item[4] for item in self.invoice_items)
        self.total_label.config(text=f"Ksh{total:.2f}")

    def save_and_print_invoice(self):
        supplier = self.supplier_entry.get()
        if not supplier or not self.invoice_items:
            messagebox.showwarning("Warning", "Enter supplier name and add items to the invoice")
            return
        
        invoice_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        for item_id, name, quantity, price, _ in self.invoice_items:
            c.execute("INSERT INTO invoices (supplier, item_id, quantity, price, invoice_date) VALUES (?, ?, ?, ?, ?)",
                     (supplier, item_id, quantity, price, invoice_date))
            c.execute("SELECT quantity FROM items WHERE id=?", (item_id,))
            current_qty = c.fetchone()[0]
            curr_num, curr_unit = self.app.parse_quantity(current_qty)
            new_num = curr_num + float(self.app.parse_quantity(quantity)[0])
            new_quantity = f"{new_num} {curr_unit}"
            c.execute("UPDATE items SET quantity=?, last_updated=? WHERE id=?", (new_quantity, invoice_date, item_id))
        conn.commit()
        conn.close()
        
        filename = f"invoice_{self.section}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph(f"Invoice for {self.section.capitalize()}", styles['Title']))
        elements.append(Paragraph(f"Supplier: {supplier}", styles['Normal']))
        elements.append(Paragraph(f"Date: {invoice_date}", styles['Normal']))
        elements.append(Paragraph("<br/><br/>", styles['Normal']))
        
        data = [["Name", "Quantity", "Price (Ksh)", "Total (Ksh)"]] + [[item[1], item[2], item[3], item[4]] for item in self.invoice_items]
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        elements.append(Paragraph(f"<br/>Total Amount: Ksh{sum(item[4] for item in self.invoice_items):.2f}", styles['Normal']))
        
        doc.build(elements)
        messagebox.showinfo("Success", f"Invoice saved as {filename}")
        webbrowser.open(filename)
        
        self.load_items(self.store_tree)
        self.invoice_items = []
        self.invoice_tree.delete(*self.invoice_tree.get_children())
        self.update_invoice_total()

    def get_categories(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT DISTINCT category FROM items WHERE section=? AND category IS NOT NULL", (self.section,))
        categories = [row[0] for row in c.fetchall()]
        conn.close()
        return categories or ["General"]

    def get_item_names(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name FROM items WHERE section=?", (self.section,))
        names = [row[0] for row in c.fetchall()]
        conn.close()
        return names

    def add_store_item(self):
        name = self.store_name.get().strip()
        category = self.store_category.get().strip() or "General"
        purchase_unit = self.store_purchase_unit.get().strip()
        quantity = self.store_quantity.get().strip()
        reorder = self.store_reorder.get().strip()
        expiry_date = self.store_expiry.get().strip()
        price = self.store_price.get().strip()
        last_updated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not name or not purchase_unit or not quantity:
            messagebox.showerror("Error", "Name, Purchase Unit, and Quantity are required!")
            return

        # Check if item exists
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT id FROM items WHERE name=? AND section=?", (name, self.section))
        existing_item = c.fetchone()

        if existing_item:
            response = messagebox.askyesno("Item Exists", f"Item '{name}' already exists in the store. Would you like to update it instead?")
            conn.close()
            if response:
                self.update_store_items()
                return
            else:
                messagebox.showinfo("Info", "Add cancelled. Use a different name or update the existing item.")
                return

        # Proceed with adding new item
        try:
            qty_num = float(quantity)
            pu_num, pu_unit = self.app.parse_quantity(purchase_unit)
            total_qty = qty_num * pu_num
            total_quantity_str = f"{total_qty} {pu_unit}"

            reorder_num = float(reorder) if reorder else 0
            reorder_qty = reorder_num * pu_num
            reorder_str = f"{reorder_qty} {pu_unit}"

            price_num = float(price) if price else 0.0
            if price_num < 0:
                messagebox.showerror("Error", "Price cannot be negative!")
                return
        except ValueError:
            messagebox.showerror("Error", "Invalid quantity, purchase unit, or price format!")
            conn.close()
            return

        c.execute("""
            INSERT INTO items (name, category, quantity, reorder_level, expiry_date, section, last_updated, purchase_unit)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (name, category, total_quantity_str, reorder_str, expiry_date, self.section, last_updated, purchase_unit))
        item_id = c.lastrowid

        if price_num > 0:
            c.execute("""
                INSERT INTO prices (item_id, price, last_updated) 
                VALUES (?, ?, ?)
            """, (item_id, price_num, last_updated))

        details = f"Added {name} with {quantity} x {purchase_unit} = {total_quantity_str}"
        if price_num > 0:
            details += f", Price Ksh{price_num:.2f} per {purchase_unit}"
        c.execute("INSERT INTO history (action, item_id, details, timestamp) VALUES (?, ?, ?, ?)",
                  ("Add", item_id, details, last_updated))

        conn.commit()
        conn.close()

        self.load_items(self.store_tree)
        self.app.load_history()
        self.check_reorder()
        self.check_expiry()
        self.store_category['values'] = self.get_categories()
        self.clear_store_entries()

    def update_store_items(self):
        name = self.store_name.get().strip()
        category = self.store_category.get().strip() or "General"
        purchase_unit = self.store_purchase_unit.get().strip()
        quantity = self.store_quantity.get().strip()
        reorder = self.store_reorder.get().strip()
        expiry_date = self.store_expiry.get().strip()
        price = self.store_price.get().strip()

        if not name or not purchase_unit or not quantity:
            messagebox.showerror("Error", "Name, Purchase Unit, and Quantity are required to update an item!")
            return

        # Validate inputs
        try:
            qty_num = float(quantity)
            pu_num, pu_unit = self.app.parse_quantity(purchase_unit)
            total_qty = qty_num * pu_num
            total_quantity_str = f"{total_qty} {pu_unit}"

            reorder_num = float(reorder) if reorder else 0
            reorder_qty = reorder_num * pu_num
            reorder_str = f"{reorder_qty} {pu_unit}"

            price_num = float(price) if price else None
            if price_num is not None and price_num < 0:
                messagebox.showerror("Error", "Price cannot be negative!")
                return
        except ValueError:
            messagebox.showerror("Error", "Invalid quantity, purchase unit, or price format!")
            return

        # Validate expiry date (optional)
        if expiry_date:
            try:
                datetime.strptime(expiry_date, "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("Error", "Invalid expiry date format (YYYY-MM-DD)!")
                return

        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()

        # Check if item exists
        c.execute("SELECT id, quantity FROM items WHERE name=? AND section=?", (name, self.section))
        item = c.fetchone()

        if not item:
            messagebox.showerror("Error", f"Item '{name}' not found in store section!")
            conn.close()
            return

        item_id, current_qty = item
        curr_num, curr_unit = self.app.parse_quantity(current_qty)
        if curr_num is None:
            messagebox.showerror("Error", f"Current quantity for {name} is invalid!")
            conn.close()
            return

        # Update quantity by adding to existing
        new_total = curr_num + total_qty
        updated_quantity_str = f"{new_total} {curr_unit or pu_unit}"

        last_updated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        c.execute("""
            UPDATE items SET name=?, category=?, quantity=?, reorder_level=?, expiry_date=?, 
            purchase_unit=?, last_updated=? WHERE id=?
        """, (name, category, updated_quantity_str, reorder_str, expiry_date or None, pu_unit, last_updated, item_id))

        if price_num is not None:
            c.execute("INSERT OR REPLACE INTO prices (item_id, price, last_updated) VALUES (?, ?, ?)",
                      (item_id, price_num, last_updated))

        details = f"Added {quantity} x {purchase_unit} to {current_qty}, New Total: {updated_quantity_str}"
        if price_num is not None:
            details += f", Price Ksh{price_num:.2f} per {purchase_unit}"
        c.execute("INSERT INTO history (action, item_id, details, timestamp) VALUES (?, ?, ?, ?)",
                  ("Update Qty", item_id, details, last_updated))

        conn.commit()
        conn.close()

        self.load_items(self.store_tree)
        self.app.load_history()
        self.check_reorder()
        self.check_expiry()
        messagebox.showinfo("Success", f"Item '{name}' updated successfully!")
        self.clear_store_entries()

    def clear_store_entries(self):
        """Clear all input fields in the store tab."""
        self.store_name.delete(0, tk.END)
        self.store_category.set("")
        self.store_purchase_unit.delete(0, tk.END)
        self.store_quantity.delete(0, tk.END)
        self.store_reorder.delete(0, tk.END)
        self.store_expiry.delete(0, tk.END)
        self.store_price.delete(0, tk.END)

    def delete_store_item(self):
        selected = self.store_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Select an item to delete")
            return
        if messagebox.askyesno("Confirm", "Are you sure you want to delete this item?"):
            item_id = self.store_tree.item(selected)["values"][0]
            name = self.store_tree.item(selected)["values"][1]
            conn = sqlite3.connect(self.app.get_db_path())
            c = conn.cursor()
            c.execute("DELETE FROM items WHERE id=?", (item_id,))
            c.execute("INSERT INTO history (action, item_id, details, timestamp) VALUES (?, ?, ?, ?)",
                     ("Delete", item_id, f"Deleted {name}", datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            conn.commit()
            conn.close()
            self.load_items(self.store_tree)
            self.app.load_history()

    def issue_store_item(self):
        selected = self.store_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Select an item to issue")
            return

        item_id = self.store_tree.item(selected)["values"][0]
        person_name = self.issue_person.get()
        quantity = self.issue_quantity.get()  # Number of purchase units
        issue_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not person_name or not quantity:
            messagebox.showerror("Error", "Person name and quantity are required")
            return

        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name, quantity, purchase_unit FROM items WHERE id=?", (item_id,))
        item = c.fetchone()
        if not item:
            conn.close()
            return

        name, current_qty, purchase_unit = item
        qty_num = float(quantity)
        pu_num, pu_unit = self.app.parse_quantity(purchase_unit)
        total_qty_issued = qty_num * pu_num
        curr_num, curr_unit = self.app.parse_quantity(current_qty)
        if total_qty_issued > curr_num:
            messagebox.showerror("Error", "Insufficient quantity in stock")
            conn.close()
            return

        new_qty = curr_num - total_qty_issued
        total_qty_str = f"{new_qty} {curr_unit}"

        c.execute("UPDATE items SET quantity=?, last_updated=? WHERE id=?", (total_qty_str, issue_date, item_id))
        c.execute("INSERT INTO issuance (item_id, person_name, quantity_issued, issue_date) VALUES (?, ?, ?, ?)",
                 (item_id, person_name, f"{quantity} x {purchase_unit}", issue_date))
        
        # Fetch price for logging
        c.execute("SELECT price FROM prices WHERE item_id=?", (item_id,))
        price = c.fetchone()
        price_per_unit = price[0] if price else 0.0
        details = f"Issued {quantity} x {purchase_unit} of {name} to {person_name}"
        if price_per_unit > 0:
            details += f", Value Ksh{price_per_unit * qty_num:.2f}"
        c.execute("INSERT INTO history (action, item_id, details, timestamp) VALUES (?, ?, ?, ?)",
                 ("Issue", item_id, details, issue_date))
        conn.commit()
        conn.close()

        self.load_items(self.store_tree)
        self.app.load_history()
        self.check_reorder()
        self.clear_issue_entries()

    def load_items(self, tree):
        tree.delete(*tree.get_children())
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("""
            SELECT i.id, i.name, i.category, i.quantity, i.reorder_level, 
                   i.expiry_date, i.last_updated, p.price, i.purchase_unit 
            FROM items i 
            LEFT JOIN prices p ON i.id = p.item_id 
            WHERE i.section=?
        """, (self.section,))
        for row in c.fetchall():
            item_id, name, category, total_qty, reorder, expiry, last_updated, price, purchase_unit = row  # 9 variables for 9 columns
            if purchase_unit and purchase_unit != 'unit':  # Only convert if purchase_unit is meaningful
                total_num, total_unit = self.app.parse_quantity(total_qty)
                pu_num, _ = self.app.parse_quantity(purchase_unit)
                num_units = total_num / pu_num if pu_num else total_num
                display_qty = f"{num_units:.2f} x {purchase_unit} ({total_qty})"
            else:
                display_qty = total_qty
            tree.insert("", "end", values=(item_id, name, category, display_qty, reorder, expiry, last_updated, f"Ksh{price:.2f}" if price is not None else "N/A"))
        conn.close()

    def search_items(self):
        search_term = self.store_search.get().lower()
        filter_type = self.search_filter.get()
        self.store_tree.delete(*self.store_tree.get_children())
        
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        
        # Define the base query and modify based on filter
        if filter_type == "Name":
            c.execute("SELECT id, name, category, quantity, reorder_level, expiry_date, last_updated FROM items WHERE section=? AND name LIKE ?",
                     (self.section, f"%{search_term}%"))
        elif filter_type == "Category":
            c.execute("SELECT id, name, category, quantity, reorder_level, expiry_date, last_updated FROM items WHERE section=? AND category LIKE ?",
                     (self.section, f"%{search_term}%"))
        elif filter_type == "ID":
            c.execute("SELECT id, name, category, quantity, reorder_level, expiry_date, last_updated FROM items WHERE section=? AND CAST(id AS TEXT) LIKE ?",
                     (self.section, f"%{search_term}%"))
        elif filter_type == "Quantity":
            c.execute("SELECT id, name, category, quantity, reorder_level, expiry_date, last_updated FROM items WHERE section=? AND CAST(quantity AS TEXT) LIKE ?",
                     (self.section, f"%{search_term}%"))
        
        # Insert results into Treeview
        for row in c.fetchall():
            self.store_tree.insert("", "end", values=row)
        
        conn.close()

    def select_item(self, event):
        selected = self.store_tree.selection()
        if selected:
            values = self.store_tree.item(selected)["values"]
            self.store_name.delete(0, tk.END)
            self.store_name.insert(0, values[1])
            self.store_category.set(values[2])

            # Extract number of purchase units from display quantity
            conn = sqlite3.connect(self.app.get_db_path())
            c = conn.cursor()
            c.execute("SELECT purchase_unit, quantity FROM items WHERE id=?", (values[0],))
            purchase_unit, total_qty = c.fetchone()
            conn.close()

            if purchase_unit:
                total_num, _ = self.app.parse_quantity(total_qty)
                pu_num, _ = self.app.parse_quantity(purchase_unit)
                num_units = total_num / pu_num if pu_num else total_num
                self.store_quantity.delete(0, tk.END)
                self.store_quantity.insert(0, f"{num_units}")
                self.store_purchase_unit.delete(0, tk.END)
                self.store_purchase_unit.insert(0, purchase_unit)
            else:
                self.store_quantity.delete(0, tk.END)
                self.store_quantity.insert(0, values[3])  # Fallback to total qty if no purchase unit
                self.store_purchase_unit.delete(0, tk.END)

            self.store_reorder.delete(0, tk.END)
            self.store_reorder.insert(0, values[4].split()[0] if values[4] and purchase_unit else values[4])  # Show only number if purchase unit exists
            self.store_expiry.delete(0, tk.END)
            self.store_expiry.insert(0, values[5] if values[5] else "")

    def check_reorder(self):
        if not self.app.notification_enabled:
            return
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name, quantity, reorder_level, purchase_unit FROM items WHERE section=?", (self.section,))
        items = c.fetchall()
        conn.close()

        for name, qty, reorder, purchase_unit in items:
            qty_num, qty_unit = self.app.parse_quantity(qty)
            reorder_num, reorder_unit = self.app.parse_quantity(reorder)
            pu_num, _ = self.app.parse_quantity(purchase_unit) if purchase_unit else (1, qty_unit)

            # Convert to number of purchase units
            current_units = qty_num / pu_num if pu_num else qty_num
            reorder_units = reorder_num / pu_num if pu_num else reorder_num

            if current_units <= reorder_units:
                notification.notify(
                    title=f"{self.section.capitalize()} Reorder Alert",
                    message=f"{name} is below reorder level ({current_units:.2f} units <= {reorder_units:.2f} units of {purchase_unit})",
                    timeout=10
                )

    def check_expiry(self):
        if not self.app.notification_enabled:
            return
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name, expiry_date FROM items WHERE section=? AND expiry_date IS NOT NULL", (self.section,))
        items = c.fetchall()
        conn.close()
        
        now = datetime.now()
        for name, expiry_date in items:
            try:
                expiry = datetime.strptime(expiry_date, "%Y-%m-%d")
                days_left = (expiry - now).days
                if days_left <= 0:
                    notification.notify(
                        title=f"{self.section.capitalize()} Expiry Alert",
                        message=f"{name} has expired on {expiry_date}",
                        timeout=10
                    )
                elif days_left <= 30:
                    notification.notify(
                        title=f"{self.section.capitalize()} Expiry Alert",
                        message=f"{name} expires in {days_left} days ({expiry_date})",
                        timeout=10
                    )
            except ValueError:
                continue
        self.app.root.after(86400000, self.check_expiry)

    def clear_store_entries(self):
        self.store_name.delete(0, tk.END)
        self.store_quantity.delete(0, tk.END)
        self.store_reorder.delete(0, tk.END)
        self.store_expiry.delete(0, tk.END)
        self.store_purchase_unit.delete(0, tk.END)
        self.store_price.delete(0, tk.END)
        self.store_category.set("")

    def clear_issue_entries(self):
        self.issue_person.delete(0, tk.END)
        self.issue_item.set("")
        self.issue_quantity.delete(0, tk.END)

    def export_to_csv(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name, category, quantity, reorder_level, expiry_date, last_updated FROM items WHERE section=?", (self.section,))
        rows = c.fetchall()
        conn.close()
        
        filename = f"{self.section}_inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        with open(filename, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Name", "Category", "Quantity", "Reorder Level", "Expiry Date", "Last Updated"])
            writer.writerows(rows)
        messagebox.showinfo("Success", f"Exported to {filename}")

    def export_to_excel(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name, category, quantity, reorder_level, expiry_date, last_updated FROM items WHERE section=?", (self.section,))
        rows = c.fetchall()
        conn.close()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{self.section.capitalize()} Inventory"
        
        headers = ["Name", "Category", "Quantity", "Reorder Level", "Expiry Date", "Last Updated"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        filename = f"{self.section}_inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        messagebox.showinfo("Success", f"Exported to {filename}")

    def export_to_pdf(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name, category, quantity, reorder_level, expiry_date, last_updated FROM items WHERE section=?", (self.section,))
        rows = c.fetchall()
        conn.close()
        
        filename = f"{self.section}_inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []
        data = [["Name", "Category", "Quantity", "Reorder Level", "Expiry Date", "Last Updated"]] + list(rows)
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), '#808080'),
            ('TEXTCOLOR', (0, 0), (-1, 0), '#FFFFFF'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), '#F5F5DC'),
            ('GRID', (0, 0), (-1, -1), 1, '#000000'),
        ]))
        elements.append(table)
        
        doc.build(elements)
        messagebox.showinfo("Success", f"Exported to {filename}")  

    def show_graph(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name, quantity, reorder_level FROM items WHERE section=?", (self.section,))
        data = c.fetchall()
        conn.close()
        
        names, quantities, reorders = [], [], []
        for name, qty, reorder in data:
            qty_num, _ = self.app.parse_quantity(qty)
            reorder_num, _ = self.app.parse_quantity(reorder)
            names.append(name)
            quantities.append(qty_num)
            reorders.append(reorder_num)
        
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.bar(names, quantities, label="Quantity", color="skyblue")
        ax.plot(names, reorders, "r--", label="Reorder Level")
        ax.set_title(f"{self.section.capitalize()} Inventory Levels")
        ax.set_xlabel("Items")
        ax.set_ylabel("Quantity (numeric only)")
        ax.legend()
        plt.xticks(rotation=45, ha="right")
        
        graph_window = tk.Toplevel(self.app.root)
        graph_window.title(f"{self.section.capitalize()} Inventory Graph")
        canvas = FigureCanvasTkAgg(fig, master=graph_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=1)

    def show_stock_summary(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        # Use aggregate functions to get accurate totals directly from SQL
        c.execute("""
            SELECT COUNT(*), 
                   SUM(CAST(SUBSTR(quantity, 1, INSTR(quantity, ' ')-1) AS REAL)), 
                   MIN(CAST(SUBSTR(quantity, 1, INSTR(quantity, ' ')-1) AS REAL)), 
                   MAX(CAST(SUBSTR(quantity, 1, INSTR(quantity, ' ')-1) AS REAL)) 
            FROM items WHERE section=?
        """, (self.section,))
        total_items, total_qty, min_qty, max_qty = c.fetchone()
        total_items = total_items or 0
        total_qty = total_qty or 0
        min_qty = min_qty or 0
        max_qty = max_qty or 0
        avg_qty = total_qty / total_items if total_items > 0 else 0

        c.execute("""
            SELECT COUNT(*) 
            FROM items 
            WHERE section=? 
            AND CAST(SUBSTR(quantity, 1, INSTR(quantity, ' ')-1) AS REAL) <= CAST(SUBSTR(reorder_level, 1, INSTR(reorder_level, ' ')-1) AS REAL)
        """, (self.section,))
        low_stock = c.fetchone()[0] or 0
        conn.close()

        report = f"{self.section.capitalize()} Stock Summary:\n\n"
        report += f"Total Items: {total_items}\n"
        report += f"Total Quantity (numeric only): {total_qty}\n"
        report += f"Average Quantity (numeric only): {avg_qty:.2f}\n"
        report += f"Minimum Quantity (numeric only): {min_qty}\n"
        report += f"Maximum Quantity (numeric only): {max_qty}\n"
        report += f"Items Below Reorder Level: {low_stock}\n"

        # Create scrollable window with Treeview
        summary_window = tk.Toplevel(self.app.root)
        summary_window.title(f"{self.section.capitalize()} Stock Summary")
        summary_window.geometry("600x400")

        tree = ttk.Treeview(summary_window, columns=("Metric", "Value"), show="headings")
        tree.heading("Metric", text="Metric")
        tree.heading("Value", text="Value")
        tree.column("Metric", width=200)
        tree.column("Value", width=200)

        metrics = [
            ("Total Items", total_items),
            ("Total Quantity (numeric only)", total_qty),
            ("Average Quantity (numeric only)", f"{avg_qty:.2f}"),
            ("Minimum Quantity (numeric only)", min_qty),
            ("Maximum Quantity (numeric only)", max_qty),
            ("Items Below Reorder Level", low_stock)
        ]
        for metric, value in metrics:
            tree.insert("", "end", values=(metric, value))

        scrollbar = ttk.Scrollbar(summary_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Button(summary_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)
    

    def low_stock_report(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name, category, quantity, reorder_level FROM items WHERE section=?", (self.section,))
        items = c.fetchall()
        conn.close()

        report = f"{self.section.capitalize()} Low Stock Report:\n\n"
        low_items = []
        for name, category, qty, reorder in items:
            qty_num, qty_unit = self.app.parse_quantity(qty)
            reorder_num, reorder_unit = self.app.parse_quantity(reorder)
            qty_converted = self.app.convert_units(qty_num, qty_unit, reorder_unit)
            if qty_converted is not None and reorder_num is not None and qty_converted <= reorder_num:
                low_items.append((name, category, qty, reorder))
                report += f"Name: {name}, Category: {category}, Quantity: {qty}, Reorder Level: {reorder}\n"

        if not low_items:
            report += "No items below reorder level.\n"

        # Create scrollable window with Treeview
        report_window = tk.Toplevel(self.app.root)
        report_window.title(f"{self.section.capitalize()} Low Stock Report")
        report_window.geometry("600x400")

        tree = ttk.Treeview(report_window, columns=("Name", "Category", "Quantity", "Reorder Level"), show="headings")
        tree.heading("Name", text="Name")
        tree.heading("Category", text="Category")
        tree.heading("Quantity", text="Quantity")
        tree.heading("Reorder Level", text="Reorder Level")
        tree.column("Name", width=150)
        tree.column("Category", width=100)
        tree.column("Quantity", width=100)
        tree.column("Reorder Level", width=100)

        if low_items:
            for item in low_items:
                tree.insert("", "end", values=item)
        else:
            tree.insert("", "end", values=("No items below reorder level", "", "", ""))

        scrollbar = ttk.Scrollbar(report_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Button(report_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)

        
    def category_analysis(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("""
            SELECT category, COUNT(*), SUM(CAST(SUBSTR(quantity, 1, INSTR(quantity, ' ')-1) AS REAL)) 
            FROM items 
            WHERE section=? 
            GROUP BY category
        """, (self.section,))
        data = c.fetchall()
        conn.close()

        report = f"{self.section.capitalize()} Category Analysis:\n\n"
        categories = []
        for cat, count, total_qty in data:
            report += f"Category: {cat}, Items: {count}, Total Quantity: {total_qty}\n"
            categories.append((cat, count, f"{total_qty}"))

        if not categories:
            report += "No categories found.\n"

        # Create scrollable window with Treeview
        analysis_window = tk.Toplevel(self.app.root)
        analysis_window.title(f"{self.section.capitalize()} Category Analysis")
        analysis_window.geometry("600x400")

        tree = ttk.Treeview(analysis_window, columns=("Category", "Items", "Total Quantity"), show="headings")
        tree.heading("Category", text="Category")
        tree.heading("Items", text="Items")
        tree.heading("Total Quantity", text="Total Quantity")
        tree.column("Category", width=200)
        tree.column("Items", width=100)
        tree.column("Total Quantity", width=150)

        if categories:
            for category in categories:
                tree.insert("", "end", values=category)
        else:
            tree.insert("", "end", values=("No categories found", "", ""))

        scrollbar = ttk.Scrollbar(analysis_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Button(analysis_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)

        # Optional: Uncomment to keep plots alongside Treeview
    
        categories, counts, quantities = zip(*data) if data else ([], [], [])
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 4))
        ax1.pie(counts, labels=categories, autopct='%1.1f%%', startangle=90)
        ax1.set_title(f"{self.section.capitalize()} Items by Category")
        ax2.bar(categories, quantities, color="lightgreen")
        ax2.set_title(f"{self.section.capitalize()} Quantity by Category")
        ax2.set_xlabel("Category")
        ax2.set_ylabel("Total Quantity (numeric only)")
        plt.xticks(rotation=45, ha="right")
        canvas = FigureCanvasTkAgg(fig, master=analysis_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=1)
        

    def issuance_trends(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        # Remove section filter; assume issuance is store-specific
        c.execute("""
            SELECT issue_date, SUM(CAST(SUBSTR(quantity_issued, 1, INSTR(quantity_issued, ' ')-1) AS REAL)) 
            FROM issuance 
            GROUP BY issue_date 
            ORDER BY issue_date
        """)
        data = c.fetchall()
        conn.close()

        report = f"{self.section.capitalize()} Issuance Trends:\n\n"
        trends = []
        for date, qty in data:
            report += f"Date: {date}, Quantity Issued: {qty}\n"
            trends.append((date, f"{qty}"))

        # Create scrollable window with Treeview
        report_window = tk.Toplevel(self.app.root)
        report_window.title(f"{self.section.capitalize()} Issuance Trends")
        report_window.geometry("600x400")

        tree = ttk.Treeview(report_window, columns=("Date", "Quantity Issued"), show="headings")
        tree.heading("Date", text="Date")
        tree.heading("Quantity Issued", text="Quantity Issued")
        tree.column("Date", width=200)
        tree.column("Quantity Issued", width=200)

        if trends:
            for trend in trends:
                tree.insert("", "end", values=trend)
        else:
            tree.insert("", "end", values=("No issuance data", ""))

        scrollbar = ttk.Scrollbar(report_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Button(report_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)

    def predictive_reorder(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("""
            SELECT i.name, i.quantity, i.reorder_level, iss.issue_date 
            FROM items i 
            LEFT JOIN issuance iss ON i.id = iss.item_id 
            WHERE i.section=?
        """, (self.section,))
        data = c.fetchall()
        conn.close()

        report = f"{self.section.capitalize()} Predictive Reorder Suggestions (ARIMA):\n\n"
        items_dict = {}
        for name, qty, reorder, issue_date in data:
            if name not in items_dict:
                items_dict[name] = {"quantity": qty, "reorder": reorder or "0", "dates": []}
            if issue_date:
                try:
                    items_dict[name]["dates"].append(datetime.strptime(issue_date, "%Y-%m-%d %H:%M:%S"))
                except ValueError:
                    continue

        predictions = []
        for name, info in items_dict.items():
            qty_num, qty_unit = self.app.parse_quantity(info["quantity"])
            reorder_num, reorder_unit = self.app.parse_quantity(info["reorder"])
            dates = sorted(info["dates"])

            if qty_num is None or reorder_num is None:
                report += f"{name}: Invalid quantity or reorder data, skipping prediction.\n"
                predictions.append((name, info["quantity"], info["reorder"], "N/A"))
                continue

            if len(dates) >= 5:
                start_date = min(dates)
                end_date = max(dates, default=datetime.now())
                days = (end_date - start_date).days + 1
                usage = np.zeros(days)
                for date in dates:
                    day_idx = (date - start_date).days
                    usage[day_idx] += 1  # Assumes 1 unit; adjust if quantity_issued exists

                try:
                    model = sm.tsa.ARIMA(usage, order=(1, 1, 1))
                    results = model.fit()
                    forecast = results.forecast(steps=30)
                    avg_usage = np.nanmean(forecast)  # Handle NaN in forecast
                    if np.isnan(avg_usage) or avg_usage < 0:
                        avg_usage = len(dates) / days  # Fallback
                except Exception:
                    avg_usage = len(dates) / days  # Fallback to average over period
            else:
                avg_usage = len(dates) / 30 if dates else 0

            suggested_reorder = max(reorder_num, int(avg_usage * 1.5 * self.app.convert_units(1, qty_unit, reorder_unit)))
            report += f"{name}: Current {info['quantity']}, Current Reorder {info['reorder']}, Avg Usage/Day {avg_usage:.2f}, Suggested Reorder {suggested_reorder} {reorder_unit}\n"
            predictions.append((name, info["quantity"], info["reorder"], f"{suggested_reorder} {reorder_unit}"))

        if not predictions:
            report += "No predictions available.\n"

        report_window = tk.Toplevel(self.app.root)
        report_window.title(f"{self.section.capitalize()} Predictive Reorder Suggestions")
        report_window.geometry("600x400")

        tree = ttk.Treeview(report_window, columns=("Name", "Current Qty", "Reorder Level", "Suggested"), show="headings")
        tree.heading("Name", text="Name")
        tree.heading("Current Qty", text="Current Quantity")
        tree.heading("Reorder Level", text="Reorder Level")
        tree.heading("Suggested", text="Suggested Reorder")
        tree.column("Name", width=150)
        tree.column("Current Qty", width=100)
        tree.column("Reorder Level", width=100)
        tree.column("Suggested", width=150)

        for pred in predictions:
            tree.insert("", "end", values=pred)
        if not predictions:
            tree.insert("", "end", values=("No predictions available", "", "", ""))

        scrollbar = ttk.Scrollbar(report_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Button(report_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)
        

    def student_feeding_program(self):
        feeding_window = tk.Toplevel(self.app.root)
        feeding_window.title("Weekly Student Feeding Program")
        feeding_window.geometry("1000x700")

        # Configure grid weights for the Toplevel window
        feeding_window.columnconfigure(0, weight=1)
        feeding_window.rowconfigure(2, weight=1) # Weight for the results frame

        # Input Frame
        input_frame = ttk.LabelFrame(feeding_window, text="Feeding Plan")
        input_frame.pack(fill="x", padx=10, pady=5)
        input_frame.columnconfigure(1, weight=1)
        input_frame.columnconfigure(3, weight=1)
        input_frame.columnconfigure(4, weight=1)

        ttk.Label(input_frame, text="Number of Students:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.feed_students = ttk.Entry(input_frame)
        self.feed_students.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(input_frame, text="Weekly Template:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.feed_template = ttk.Combobox(input_frame, values=["Create New"] + self.get_meal_templates())
        self.feed_template.set("Create New")
        self.feed_template.grid(row=0, column=3, padx=5, pady=5, sticky="ew")
        self.feed_template.bind("<<ComboboxSelected>>", self.load_template)

        ttk.Button(input_frame, text="Calculate Weekly Plan", command=self.calculate_feeding).grid(row=0, column=4, padx=5, pady=5, sticky="ew")

        # Template Creation Frame (shown only if "Create New" is selected)
        self.template_frame = ttk.LabelFrame(feeding_window, text="Create Weekly Template")
        self.template_frame.pack(fill="x", padx=10, pady=5)
        self.template_frame.columnconfigure(1, weight=1)
        self.template_frame.columnconfigure(3, weight=1)
        self.template_frame.columnconfigure(5, weight=1)
        self.meal_entries = {}
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        meals = ["Breakfast", "Lunch", "Dinner"]

        for i, day in enumerate(days):
            for j, meal in enumerate(meals):
                label = f"{meal} ({day})"
                ttk.Label(self.template_frame, text=label + ":").grid(row=i, column=j*2, padx=5, pady=2, sticky="w")
                entry = tk.Text(self.template_frame, height=2, width=30)
                entry.grid(row=i, column=j*2+1, padx=5, pady=2, sticky="ew")
                self.meal_entries[f"{meal}-{day}"] = entry

        ttk.Button(self.template_frame, text="Save Template", command=self.save_template).grid(row=len(days), column=0, columnspan=6, pady=5, sticky="ew")

        # Results Frame
        result_frame = ttk.LabelFrame(feeding_window, text="Weekly Feeding Plan Results")
        result_frame.pack(fill="both", expand=True, padx=10, pady=5)
        result_frame.columnconfigure(0, weight=1)
        result_frame.columnconfigure(1, weight=1)
        result_frame.rowconfigure(0, weight=1)

        self.feed_result = tk.Text(result_frame, height=15, width=70)
        self.feed_result.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

        # Assuming you are using matplotlib for the graph
        try:
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from matplotlib import pyplot as plt
            self.feed_canvas = FigureCanvasTkAgg(plt.Figure(figsize=(5, 5)), master=result_frame)
            self.feed_canvas.get_tk_widget().grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
        except ImportError:
            ttk.Label(result_frame, text="Matplotlib not installed. Cannot display graph.").grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
            self.feed_canvas = None

        # Buttons Frame
        button_frame = ttk.Frame(feeding_window)
        button_frame.pack(fill="x", padx=10, pady=5)
        button_frame.columnconfigure(0, weight=1)
        button_frame.columnconfigure(1, weight=1)
        ttk.Button(button_frame, text="Export PDF Report", command=self.export_feeding_report).grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        ttk.Button(button_frame, text="Suggest Reorders", command=self.suggest_reorders).grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        # Show/hide template frame based on selection
        self.load_template(None)
        
    def calculate_feeding(self):
        try:
            num_students = int(self.feed_students.get())
            template_name = self.feed_template.get()
            if template_name == "Create New":
                messagebox.showwarning("Warning", "Please select or save a template first")
                return

            conn = sqlite3.connect(self.app.get_db_path())
            c = conn.cursor()
            c.execute("SELECT requirements FROM meal_templates WHERE name=?", (template_name,))
            row = c.fetchone()
            conn.close()
            
            if not row:
                messagebox.showerror("Error", "Template not found")
                return

            # Parse weekly requirements
            reqs = row[0].split("|")
            weekly_reqs = {}
            for meal_req in reqs:
                meal_day, items = meal_req.split(":", 1)
                req_dict = {}
                for item_qty in items.split(";"):
                    if ":" in item_qty:
                        item, qty = item_qty.split(":")
                        req_dict[item.strip()] = qty.strip()
                weekly_reqs[meal_day] = req_dict

            result = f"Weekly Feeding Plan for {num_students} Students ({template_name}):\n\n"
            total_cost = 0
            total_calories = 0
            total_protein = 0
            weekly_data = {}

            conn = sqlite3.connect(self.app.get_db_path())
            c = conn.cursor()
            
            for meal_day, req_dict in weekly_reqs.items():
                result += f"{meal_day}:\n"
                for item, req_qty in req_dict.items():
                    req_num, req_unit = self.app.parse_quantity(req_qty)
                    total_needed = req_num * num_students
                    c.execute("""
                        SELECT i.quantity, p.price, i.calories, i.protein 
                        FROM items i 
                        LEFT JOIN prices p ON i.id = p.item_id 
                        WHERE i.section=? AND i.name=?
                    """, (self.section, item))
                    row = c.fetchone()
                    
                    if row:  # Item in inventory
                        avail_qty, price, calories, protein = row
                        avail_num, avail_unit = self.app.parse_quantity(avail_qty)
                        avail_converted = self.app.convert_units(avail_num, avail_unit, req_unit)
                        unit_cost = price if price is not None else 0.0  # Price per purchase unit
                        c.execute("SELECT purchase_unit FROM items WHERE name=? AND section=?", (item, self.section))
                        purchase_unit = c.fetchone()[0]
                        pu_num, _ = self.app.parse_quantity(purchase_unit) if purchase_unit else (1, req_unit)
                        cost = total_needed * unit_cost / pu_num if pu_num else total_needed * unit_cost
                        calories_needed = total_needed * calories
                        protein_needed = total_needed * protein
                        status = "Sufficient" if avail_converted >= total_needed else f"Shortfall {(total_needed - avail_converted):.2f} {req_unit}"
                    else:  # Daily supply
                        avail_qty = "N/A (Daily Supply)"
                        unit_cost = 0.0
                        calories = 0.0
                        protein = 0.0
                        cost = 0.0
                        calories_needed = 0.0
                        protein_needed = 0.0
                        status = "Daily Supply"

                    total_cost += cost
                    total_calories += calories_needed
                    total_protein += protein_needed
                    
                    result += f"  {item}: Needed {total_needed:.2f} {req_unit}, Available {avail_qty}, {status}, Cost Ksh{cost:.2f}, Calories {calories_needed:.2f} kcal, Protein {protein_needed:.2f} g\n"
                    
                    if item in weekly_data:
                        weekly_data[item]["total_needed"] += total_needed
                        weekly_data[item]["cost"] += cost
                    else:
                        weekly_data[item] = {"total_needed": total_needed, "available": avail_converted if row else 0, "cost": cost}
            
            conn.close()
            result += f"\nWeekly Totals:\n"
            result += f"Total Cost: Ksh{total_cost:.2f}\n"
            result += f"Per Student Per Day: Calories {(total_calories/num_students/7):.2f} kcal, Protein {(total_protein/num_students/7):.2f} g"
            
            self.feed_result.delete("1.0", tk.END)
            self.feed_result.insert("1.0", result)
            logging.info(f"Calculated weekly feeding plan for {num_students} students using {template_name}")

            # Visualization: Pie Chart of Weekly Costs
            fig = self.feed_canvas.figure
            fig.clear()
            ax = fig.add_subplot(111)
            labels = list(weekly_data.keys())
            costs = [data["cost"] for data in weekly_data.values()]
            ax.pie(costs, labels=labels, autopct='%1.1f%%', startangle=90)
            ax.set_title("Weekly Cost Distribution")
            self.feed_canvas.draw()

        except ValueError as e:
            messagebox.showerror("Error", "Invalid input: Ensure number of students is valid")
            logging.error(f"Feeding calculation error: {str(e)}")


    def get_meal_templates(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT name FROM meal_templates")
        templates = [row[0] for row in c.fetchall()]
        conn.close()
        return templates

    def load_template(self, event):
        template_name = self.feed_template.get()
        self.template_frame.pack_forget() if template_name != "Create New" else self.template_frame.pack(fill="x", padx=10, pady=5)
        
        if template_name != "Create New":
            conn = sqlite3.connect(self.app.get_db_path())
            c = conn.cursor()
            c.execute("SELECT requirements FROM meal_templates WHERE name=?", (template_name,))
            row = c.fetchone()
            conn.close()
            
            if row:
                reqs = row[0].split("|")
                for meal_req in reqs:
                    meal_day, items = meal_req.split(":", 1)
                    if meal_day in self.meal_entries:
                        self.meal_entries[meal_day].delete("1.0", tk.END)
                        self.meal_entries[meal_day].insert("1.0", items.replace(";", "\n"))
                logging.info(f"Loaded weekly template: {template_name}")

    def save_template(self):
        template_name = tk.simpledialog.askstring("Save Template", "Enter template name:")
        if not template_name:
            return
        
        requirements = []
        for meal_day, entry in self.meal_entries.items():
            items = entry.get("1.0", tk.END).strip().replace("\n", ";")
            if items:
                requirements.append(f"{meal_day}:{items}")
        requirements_str = "|".join(requirements)
        
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("INSERT OR REPLACE INTO meal_templates (name, requirements) VALUES (?, ?)", (template_name, requirements_str))
        conn.commit()
        conn.close()
        
        self.feed_template['values'] = ["Create New"] + self.get_meal_templates()
        self.feed_template.set(template_name)
        messagebox.showinfo("Success", f"Weekly template '{template_name}' saved")
        logging.info(f"Saved weekly template: {template_name}")

    def export_feeding_report(self):
        num_students = self.feed_students.get()
        template_name = self.feed_template.get()
        if not num_students or template_name == "Create New":
            messagebox.showwarning("Warning", "Calculate a weekly plan first")
            return
        
        filename = f"weekly_feeding_plan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.append(Paragraph(f"Weekly Feeding Plan for {num_students} Students ({template_name})", styles['Title']))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        elements.append(Paragraph("<br/>", styles['Normal']))
        
        data = [["Meal/Day", "Item", "Needed", "Available", "Status", "Cost (Ksh)", "Calories (kcal)", "Protein (g)"]]
        current_meal = ""
        for line in self.feed_result.get("1.0", tk.END).split("\n"):
            if line.strip() and ":" in line and "Weekly Totals" not in line and "Per Student" not in line:
                if not line.startswith("  "):
                    current_meal = line.strip()
                else:
                    parts = line.split(", ")
                    item = parts[0].split(":")[0].strip()
                    needed = parts[0].split(":")[1].split("Needed")[1].strip()
                    avail = parts[1].split("Available")[1].strip()
                    status = parts[2].strip()
                    cost = parts[3].split("Ksh")[1].strip()
                    calories = parts[4].split("Calories")[1].strip()
                    protein = parts[5].split("Protein")[1].strip()
                    data.append([current_meal if item == parts[0].split(":")[0].strip() else "", item, needed, avail, status, cost, calories, protein])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('SPAN', (0, 1), (0, -1), 'LEFT'),  # Merge meal/day cells where empty
        ]))
        elements.append(table)
        
        total_text = self.feed_result.get("1.0", tk.END).split("\n")[-3:-1]
        for line in total_text:
            elements.append(Paragraph(line, styles['Normal']))
        
        doc.build(elements)
        messagebox.showinfo("Success", f"Report saved as {filename}")
        webbrowser.open(filename)
        logging.info(f"Exported weekly feeding plan to {filename}")

    def suggest_reorders(self):
        num_students = int(self.feed_students.get() or 0)
        template_name = self.feed_template.get()
        if template_name == "Create New":
            messagebox.showwarning("Warning", "Calculate a weekly plan first")
            return

        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT requirements FROM meal_templates WHERE name=?", (template_name,))
        row = c.fetchone()
        conn.close()

        if not row:
            messagebox.showwarning("Warning", f"Template '{template_name}' not found!")
            return

        weekly_reqs = {}
        for meal_req in row[0].split("|"):
            if ":" not in meal_req:
                continue
            _, items = meal_req.split(":", 1)
            req_dict = {item.split(":")[0].strip(): item.split(":")[1].strip() for item in items.split(";") if ":" in item}
            for item, qty in req_dict.items():
                qty_num, qty_unit = self.app.parse_quantity(qty)
                if item in weekly_reqs:
                    weekly_reqs[item]["qty"] += qty_num
                else:
                    weekly_reqs[item] = {"qty": qty_num, "unit": qty_unit}

        report = f"{self.section.capitalize()} Weekly Reorder Suggestions:\n\n"
        suggestions = []
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()

        for item, data in weekly_reqs.items():
            total_needed = data["qty"] * num_students * 7  # 7 days
            req_unit = data["unit"]
            c.execute("SELECT quantity, unit_cost, reorder_level FROM items WHERE section=? AND name=?", (self.section, item))
            row = c.fetchone()

            if row:
                avail_qty, unit_cost, reorder = row
                avail_num, avail_unit = self.app.parse_quantity(avail_qty)
                reorder_num, reorder_unit = self.app.parse_quantity(reorder)
                avail_converted = self.app.convert_units(avail_num, avail_unit, req_unit)
                if avail_converted < total_needed:
                    shortfall = total_needed - avail_converted
                    safety_stock = total_needed * 0.2
                    order_qty = shortfall + safety_stock
                    cost = order_qty * (unit_cost or 0)
                    report += f"{item}: Shortfall {shortfall:.2f} {req_unit}, Suggested Order {order_qty:.2f} {req_unit} (Cost Ksh{cost:.2f})\n"
                    suggestions.append((item, f"{shortfall:.2f} {req_unit}", f"{order_qty:.2f} {req_unit}", f"Ksh{cost:.2f}"))
                else:
                    suggestions.append((item, "0", "0", "Ksh0.00"))
            else:
                report += f"{item}: Not in inventory, Order {total_needed:.2f} {req_unit}\n"
                suggestions.append((item, "Not in inventory", f"{total_needed:.2f} {req_unit}", "N/A"))

        conn.close()
        if not suggestions:
            report += "No reorders needed.\n"

        # Create scrollable window with Treeview
        report_window = tk.Toplevel(self.app.root)
        report_window.title(f"{self.section.capitalize()} Weekly Reorder Suggestions")
        report_window.geometry("600x400")

        tree = ttk.Treeview(report_window, columns=("Item", "Shortfall", "Suggested Order", "Cost"), show="headings")
        tree.heading("Item", text="Item")
        tree.heading("Shortfall", text="Shortfall")
        tree.heading("Suggested Order", text="Suggested Order")
        tree.heading("Cost", text="Cost")
        tree.column("Item", width=150)
        tree.column("Shortfall", width=100)
        tree.column("Suggested Order", width=150)
        tree.column("Cost", width=100)

        for suggestion in suggestions:
            tree.insert("", "end", values=suggestion)

        scrollbar = ttk.Scrollbar(report_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Button(report_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)
        logging.info("Generated weekly reorder suggestions")


    def audit_report(self):
        # Fetch all data from the database
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT id, name, category, quantity, reorder_level, expiry_date, last_updated, section, purchase_unit FROM items WHERE section=?", (self.section,))
        items = c.fetchall()
        c.execute("SELECT i.name, p.price, p.last_updated FROM items i LEFT JOIN prices p ON i.id = p.item_id WHERE i.section=?", (self.section,))
        prices = c.fetchall()
        c.execute("SELECT inv.supplier, i.name, inv.quantity, inv.price, inv.invoice_date FROM invoices inv JOIN items i ON inv.item_id = i.id WHERE i.section=?", (self.section,))
        invoices = c.fetchall()
        c.execute("SELECT i.name, iss.person_name, iss.quantity_issued, iss.issue_date FROM issuance iss JOIN items i ON iss.item_id = i.id WHERE i.section=?", (self.section,))
        issuances = c.fetchall()
        c.execute("SELECT student_id, student_name, item_name, report_date, status FROM broken_items WHERE item_id IN (SELECT id FROM items WHERE section=?)", (self.section,))
        broken_items = c.fetchall()
        c.execute("SELECT action, i.name, details, timestamp FROM history h LEFT JOIN items i ON h.item_id = i.id WHERE i.section=?", (self.section,))
        history = c.fetchall()
        c.execute("SELECT id, username FROM users")
        users = c.fetchall()
        c.execute("SELECT subject, form, num_students, topic, subtopic, time, status FROM practical_reports")
        practical_reports = c.fetchall()
        c.execute("SELECT name, requirements FROM meal_templates")
        meal_templates = c.fetchall()
        c.execute("SELECT i.name, b.batch_number, b.quantity, b.unit_cost, b.received_date, b.expiry_date FROM batches b JOIN items i ON b.item_id = i.id WHERE i.section=?", (self.section,))
        batches = c.fetchall()
        conn.close()

        # Comprehensive report for PDF
        report_title = f"Comprehensive {self.section.capitalize()} Audit Report - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        report = f"{report_title}\n\n"

        # Items Section
        report += "Items:\n"
        if not items:
            report += "No items found.\n"
        for item in items:
            report += f"ID: {item[0]}, Name: {item[1]}, Category: {item[2]}, Quantity: {item[3]}, Reorder: {item[4]}, Expiry: {item[5]}, Updated: {item[6]}, Section: {item[7]}, Purchase Unit: {item[8]}\n"

        # Prices Section
        report += "\nPrices:\n"
        if not prices:
            report += "No price data available.\n"
        for price in prices:
            price_value = f"Ksh{float(price[1]):.2f}" if price[1] is not None else "KshN/A"
            report += f"Name: {price[0]}, Price: {price_value}, Updated: {price[2] or 'N/A'}\n"

        # Invoices Section
        report += "\nInvoices:\n"
        if not invoices:
            report += "No invoices found.\n"
        for inv in invoices:
            inv_price = f"Ksh{float(inv[3]):.2f}" if inv[3] is not None else "KshN/A"
            report += f"Supplier: {inv[0]}, Item: {inv[1]}, Quantity: {inv[2]}, Price: {inv_price}, Date: {inv[4] or 'N/A'}\n"

        # Issuance Section
        report += "\nIssuance:\n"
        if not issuances:
            report += "No issuance records found.\n"
        for iss in issuances:
            report += f"Item: {iss[0]}, Person: {iss[1]}, Quantity Issued: {iss[2]}, Date: {iss[3] or 'N/A'}\n"

        # Broken Items Section
        report += "\nBroken Items:\n"
        if not broken_items:
            report += "No broken items reported.\n"
        for broken in broken_items:
            report += f"Student ID: {broken[0]}, Student Name: {broken[1]}, Item: {broken[2]}, Report Date: {broken[3]}, Status: {broken[4]}\n"

        # History Section
        report += "\nHistory:\n"
        if not history:
            report += "No history records found.\n"
        for hist in history:
            report += f"Action: {hist[0]}, Item: {hist[1] or 'N/A'}, Details: {hist[2]}, Timestamp: {hist[3]}\n"

        # Users Section
        report += "\nUsers:\n"
        if not users:
            report += "No users found.\n"
        for user in users:
            report += f"ID: {user[0]}, Username: {user[1]}\n"

        # Meal Templates Section (store only)
        if self.section == "store":
            report += "\nMeal Templates:\n"
            if not meal_templates:
                report += "No meal templates found.\n"
            for meal in meal_templates:
                report += f"Name: {meal[0]}, Requirements: {meal[1]}\n"

        # Batches Section
        report += "\nBatches:\n"
        if not batches:
            report += "No batches found.\n"
        for batch in batches:
            report += f"Item: {batch[0]}, Batch Number: {batch[1]}, Quantity: {batch[2]}, Unit Cost: Ksh{float(batch[3]):.2f}, Received: {batch[4]}, Expiry: {batch[5]}\n"

        # Create Treeview window
        audit_window = tk.Toplevel(self.app.root)
        audit_window.title(f"Comprehensive {self.section.capitalize()} Audit Report")
        audit_window.geometry("1000x800")
        audit_window.columnconfigure(0, weight=1)
        audit_window.rowconfigure(0, weight=1)

        tree_frame = ttk.Frame(audit_window)
        tree_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=5)
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

        audit_tree = ttk.Treeview(tree_frame, columns=("Details",), show="tree headings")
        audit_tree.heading("Details", text="Audit Details")
        audit_tree.column("Details", stretch=tk.YES, width=900)
        audit_tree.grid(row=0, column=0, sticky="nsew")

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=audit_tree.yview)
        audit_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.grid(row=0, column=1, sticky="ns")

        # Populate Treeview
        def add_section(parent, title, data, format_func):
            section_node = audit_tree.insert(parent, "end", text=title, open=True)
            if not data:
                audit_tree.insert(section_node, "end", text="No data found.")
            else:
                for entry in data:
                    audit_tree.insert(section_node, "end", text=format_func(entry))

        format_item = lambda x: f"ID: {x[0]}, Name: {x[1]}, Category: {x[2]}, Qty: {x[3]}, Reorder: {x[4]}, Expiry: {x[5]}, Updated: {x[6]}, Section: {x[7]}, Purchase Unit: {x[8]}"
        format_price = lambda x: f"Name: {x[0]}, Price: {'Ksh' + f'{float(x[1]):.2f}' if x[1] is not None else 'KshN/A'}, Updated: {x[2] or 'N/A'}"
        format_invoice = lambda x: f"Supplier: {x[0]}, Item: {x[1]}, Qty: {x[2]}, Price: {'Ksh' + f'{float(x[3]):.2f}' if x[3] is not None else 'KshN/A'}, Date: {x[4] or 'N/A'}"
        format_issuance = lambda x: f"Item: {x[0]}, Person: {x[1]}, Qty Issued: {x[2]}, Date: {x[3] or 'N/A'}"
        format_broken = lambda x: f"Student ID: {x[0]}, Student Name: {x[1]}, Item: {x[2]}, Report Date: {x[3]}, Status: {x[4]}"
        format_history = lambda x: f"Action: {x[0]}, Item: {x[1] or 'N/A'}, Details: {x[2]}, Timestamp: {x[3]}"
        format_user = lambda x: f"ID: {x[0]}, Username: {x[1]}"
        format_meal = lambda x: f"Name: {x[0]}, Requirements: {x[1]}"
        format_batch = lambda x: f"Item: {x[0]}, Batch: {x[1]}, Qty: {x[2]}, Cost: Ksh{float(x[3]):.2f}, Received: {x[4]}, Expiry: {x[5]}"

        add_section("", "Items", items, format_item)
        add_section("", "Prices", prices, format_price)
        add_section("", "Invoices", invoices, format_invoice)
        add_section("", "Issuance", issuances, format_issuance)
        add_section("", "Broken Items", broken_items, format_broken)
        add_section("", "History", history, format_history)
        add_section("", "Users", users, format_user)
        if self.section == "store":
            add_section("", "Meal Templates", meal_templates, format_meal)
        add_section("", "Batches", batches, format_batch)

        # Download Button
        button_frame = ttk.Frame(audit_window)
        button_frame.grid(row=1, column=0, pady=10, sticky="ew")
        button_frame.columnconfigure(0, weight=1)
        ttk.Button(button_frame, text="Download PDF", command=lambda: self.download_audit_pdf(report)).grid(row=0, column=0, padx=5, pady=5)

        logging.info(f"Generated comprehensive {self.section} audit report")

    def download_audit_pdf(self, report):
        filename = f"{self.section}_audit_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Log the raw report for debugging
        logging.debug(f"Raw report content:\n{report}")

        # Split report into lines
        lines = report.split("\n")
        if not lines:
            logging.error("Report is empty, no lines to process.")
            elements.append(Paragraph("No data available.", styles['Normal']))
        else:
            # Add title
            elements.append(Paragraph(lines[0], styles['Title']))
            elements.append(Paragraph("<br/>", styles['Normal']))
            logging.debug(f"Added title: {lines[0]}")

            current_section = None
            table_data = []
            headers = None

            for line in lines[1:]:
                line = line.strip()
                if not line:
                    continue  # Skip empty lines

                # Check if this is a section header
                if line.endswith(":") and "No " not in line:
                    # Finish previous section
                    if current_section and table_data:
                        logging.debug(f"Building table for {current_section} with {len(table_data)} rows")
                        table = Table([headers] + table_data)
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ]))
                        elements.append(table)
                        elements.append(Paragraph("<br/>", styles['Normal']))
                        table_data = []
                        headers = None
                    elif current_section:
                        logging.debug(f"Adding plain text for {current_section}")
                        elements.append(Paragraph(current_section, styles['Heading2']))

                    # Start new section
                    current_section = line
                    elements.append(Paragraph(line, styles['Heading2']))
                    logging.debug(f"New section: {line}")

                # Handle "No data" messages or data rows
                elif current_section:
                    if "No " in line:
                        elements.append(Paragraph(line, styles['Normal']))
                        logging.debug(f"Added no-data line: {line}")
                    else:
                        # Parse data rows
                        if current_section == "Items:":
                            headers = ["ID", "Name", "Category", "Quantity", "Reorder", "Expiry", "Updated", "Section", "Purchase Unit"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "Prices:":
                            headers = ["Name", "Price", "Updated"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "Invoices:":
                            headers = ["Supplier", "Item", "Quantity", "Price", "Date"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "Issuance:":
                            headers = ["Item", "Person", "Quantity Issued", "Date"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "Broken Items:":
                            headers = ["Student ID", "Student Name", "Item", "Report Date", "Status"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "History:":
                            headers = ["Action", "Item", "Details", "Timestamp"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "Users:":
                            headers = ["ID", "Username"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "Meal Templates:":
                            headers = ["Name", "Requirements"]
                            parts = line.split(", ", 1)  # Only split on first comma for requirements
                            table_data.append([part.split(": ")[1] for part in parts])
                        elif current_section == "Batches:":
                            headers = ["Item", "Batch Number", "Quantity", "Unit Cost", "Received", "Expiry"]
                            parts = line.split(", ")
                            table_data.append([part.split(": ")[1] for part in parts])
                        logging.debug(f"Added row to {current_section}: {line}")

            # Handle the last section
            if current_section and table_data:
                logging.debug(f"Building final table for {current_section} with {len(table_data)} rows")
                table = Table([headers] + table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)
            elif current_section:
                logging.debug(f"Adding final plain text for {current_section}")
                elements.append(Paragraph(current_section, styles['Heading2']))

        # Build the PDF
        try:
            doc.build(elements)
            logging.info(f"PDF built successfully: {filename}")
        except Exception as e:
            logging.error(f"Failed to build PDF: {str(e)}")
            messagebox.showerror("Error", f"Failed to generate PDF: {str(e)}")
            return

        messagebox.showinfo("Success", f"Audit report saved as {filename}")
        webbrowser.open(filename)
        logging.info(f"Downloaded {self.section} audit report as {filename}")

    def manage_daily_suppliers(self):
        supplier_window = tk.Toplevel(self.app.root)
        supplier_window.title("Daily Supplier Deliveries")
        supplier_window.geometry("800x600")

        # Configure grid weights for the Toplevel window
        supplier_window.columnconfigure(1, weight=1)
        supplier_window.rowconfigure(1, weight=1)

        ttk.Label(supplier_window, text="Supplier Name:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.supplier_name = ttk.Entry(supplier_window)
        self.supplier_name.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        self.supplier_tree = ttk.Treeview(supplier_window, columns=("ID", "Name", "Quantity", "Date"), show="headings")
        self.supplier_tree.heading("ID", text="ID")
        self.supplier_tree.heading("Name", text="Name")
        self.supplier_tree.heading("Quantity", text="Quantity")
        self.supplier_tree.heading("Date", text="Date")
        for col in ("ID", "Name", "Quantity", "Date"):
            self.supplier_tree.column(col, stretch=tk.YES)
        self.supplier_tree.grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")

        ttk.Label(supplier_window, text="Item Name:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.supplier_item = ttk.Combobox(supplier_window, values=self.get_item_names())
        self.supplier_item.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(supplier_window, text="Quantity Delivered (e.g., '5 units'):").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.supplier_qty = ttk.Entry(supplier_window)
        self.supplier_qty.grid(row=3, column=1, padx=5, pady=5, sticky="ew")

        button_frame = ttk.Frame(supplier_window)
        button_frame.grid(row=4, column=0, columnspan=2, pady=10, sticky="ew")
        button_frame.columnconfigure(0, weight=1)
        button_frame.columnconfigure(1, weight=1)

        ttk.Button(button_frame, text="Add Delivery", command=self.add_supplier_delivery).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(button_frame, text="Download Report", command=self.download_supplier_report).grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        self.load_supplier_deliveries()

    def load_supplier_deliveries(self):
        self.supplier_tree.delete(*self.supplier_tree.get_children())
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT inv.id, i.name, inv.quantity, inv.invoice_date FROM invoices inv JOIN items i ON inv.item_id = i.id WHERE i.section=?", (self.section,))
        for row in c.fetchall():
            self.supplier_tree.insert("", "end", values=row)
        conn.close()

    def add_supplier_delivery(self):
        supplier = self.supplier_name.get()
        item_name = self.supplier_item.get()
        qty = self.supplier_qty.get()
        date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT id, quantity FROM items WHERE section=? AND name=?", (self.section, item_name))
        item = c.fetchone()
        if not item:
            messagebox.showerror("Error", "Item not found")
            conn.close()
            return

        item_id, current_qty = item
        curr_num, curr_unit = self.app.parse_quantity(current_qty)
        new_num, new_unit = self.app.parse_quantity(qty)
        updated_num = curr_num + self.app.convert_units(new_num, new_unit, curr_unit)
        updated_qty = f"{updated_num} {curr_unit}"

        c.execute("INSERT INTO invoices (supplier, item_id, quantity, price, invoice_date) VALUES (?, ?, ?, ?, ?)",
              (supplier, item_id, qty, 0.0, date))  # Price set to 0 as it's daily delivery
        c.execute("UPDATE items SET quantity=?, last_updated=? WHERE id=?", (updated_qty, date, item_id))
        c.execute("INSERT INTO history (action, item_id, details, timestamp) VALUES (?, ?, ?, ?)",
                ("Delivery", item_id, f"{supplier} delivered {qty} of {item_name}", date))
        conn.commit()
        conn.close()

        self.load_supplier_deliveries()
        self.load_items(self.store_tree)
        self.app.load_history()
        logging.info(f"Added delivery from {supplier} for {item_name}: {qty}")

    def download_supplier_report(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT inv.supplier, i.name, inv.quantity, inv.invoice_date FROM invoices inv JOIN items i ON inv.item_id = i.id WHERE i.section=?", (self.section,))
        rows = c.fetchall()
        conn.close()

        filename = f"supplier_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        with open(filename, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Supplier", "Item Name", "Quantity", "Date"])
            writer.writerows(rows)
        messagebox.showinfo("Success", f"Report downloaded as {filename}")
        logging.info("Downloaded supplier report")

    # Stock Valuation (Average Cost Method)
    def calculate_stock_value(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("""
            SELECT i.name, i.quantity, i.purchase_unit, p.price 
            FROM items i 
            LEFT JOIN prices p ON i.id = p.item_id 
            WHERE i.section=?
        """, (self.section,))
        items = c.fetchall()
        conn.close()

        total_value = 0
        report = f"{self.section.capitalize()} Stock Valuation (Based on Price per Purchase Unit):\n\n"
        stock_values = []
        for name, qty, purchase_unit, price in items:
            qty_num, qty_unit = self.app.parse_quantity(qty)
            pu_num, _ = self.app.parse_quantity(purchase_unit) if purchase_unit else (1, qty_unit)
            num_units = qty_num / pu_num if pu_num else qty_num
            price_per_unit = price if price else 0.0
            value = num_units * price_per_unit
            total_value += value
            report += f"{name}: {num_units:.2f} x {purchase_unit} ({qty}), Price Ksh{price_per_unit:.2f} per {purchase_unit}, Value Ksh{value:.2f}\n"
            stock_values.append((name, f"{num_units:.2f} {purchase_unit}", f"Ksh{price_per_unit:.2f}", f"Ksh{value:.2f}"))
        report += f"\nTotal Stock Value: Ksh{total_value:.2f}\n"

        # Create scrollable window with Treeview
        report_window = tk.Toplevel(self.app.root)
        report_window.title(f"{self.section.capitalize()} Stock Valuation")
        report_window.geometry("600x400")

        tree = ttk.Treeview(report_window, columns=("Name", "Units", "Price per Unit", "Value"), show="headings")
        tree.heading("Name", text="Name")
        tree.heading("Units", text="Units")
        tree.heading("Price per Unit", text="Price per Unit")
        tree.heading("Value", text="Value")
        tree.column("Name", width=150)
        tree.column("Units", width=100)
        tree.column("Price per Unit", width=100)
        tree.column("Value", width=100)

        for value in stock_values:
            tree.insert("", "end", values=value)
        tree.insert("", "end", values=("", "", "Total", f"Ksh{total_value:.2f}"))

        scrollbar = ttk.Scrollbar(report_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Button(report_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)
        logging.info("Generated stock valuation report")

    # ABC Analysis
    def abc_analysis(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("""
            SELECT i.name, i.quantity, i.purchase_unit, p.price 
            FROM items i 
            LEFT JOIN prices p ON i.id = p.item_id 
            WHERE i.section=?
        """, (self.section,))
        items = c.fetchall()
        conn.close()

        usage_values = []
        for name, qty, purchase_unit, price in items:
            qty_num, qty_unit = self.app.parse_quantity(qty)
            pu_num, _ = self.app.parse_quantity(purchase_unit) if purchase_unit else (1, qty_unit)
            num_units = qty_num / pu_num if pu_num else qty_num
            price_per_unit = price if price else 0.0
            value = num_units * price_per_unit
            usage_values.append((name, value))

        usage_values.sort(key=lambda x: x[1], reverse=True)
        total_value = sum(val for _, val in usage_values) or 1  # Avoid division by zero

        a_items, b_items, c_items = [], [], []
        cumulative = 0
        for name, value in usage_values:
            cumulative += value
            percent = cumulative / total_value
            if percent <= 0.8:
                a_items.append((name, f"Ksh{value:.2f}", "A"))
            elif percent <= 0.95:
                b_items.append((name, f"Ksh{value:.2f}", "B"))
            else:
                c_items.append((name, f"Ksh{value:.2f}", "C"))

        report = f"{self.section.capitalize()} ABC Analysis (Based on Price per Purchase Unit):\n\n"
        report += "Category A (High Value, ~80%):\n" + "".join(f"{name}: {value}\n" for name, value, _ in a_items)
        report += "\nCategory B (Medium Value, ~15%):\n" + "".join(f"{name}: {value}\n" for name, value, _ in b_items)
        report += "\nCategory C (Low Value, ~5%):\n" + "".join(f"{name}: {value}\n" for name, value, _ in c_items)

        # Create scrollable window with Treeview
        report_window = tk.Toplevel(self.app.root)
        report_window.title(f"{self.section.capitalize()} ABC Analysis")
        report_window.geometry("600x400")

        tree = ttk.Treeview(report_window, columns=("Name", "Value", "Category"), show="headings")
        tree.heading("Name", text="Name")
        tree.heading("Value", text="Value")
        tree.heading("Category", text="Category")
        tree.column("Name", width=200)
        tree.column("Value", width=100)
        tree.column("Category", width=100)

        for item in a_items + b_items + c_items:
            tree.insert("", "end", values=item)

        scrollbar = ttk.Scrollbar(report_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Button(report_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)
        logging.info("Performed ABC analysis")

    # Safety Stock Calculation
    def calculate_safety_stock(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("""
            SELECT i.name, i.quantity, i.reorder_level, COUNT(iss.id) as issues 
            FROM items i LEFT JOIN issuance iss ON i.id = iss.item_id 
            WHERE i.section=? 
            GROUP BY i.id, i.name, i.quantity, i.reorder_level
        """, (self.section,))
        items = c.fetchall()
        conn.close()

        report = f"{self.section.capitalize()} Safety Stock Calculation (Max Usage - Avg Usage):\n\n"
        safety_stocks = []
        for name, qty, reorder, issues in items:
            qty_num, qty_unit = self.app.parse_quantity(qty)
            reorder_num, _ = self.app.parse_quantity(reorder)
            avg_usage = issues / 30 if issues else 1  # Fallback: 1 unit/day
            max_usage = max(issues, avg_usage * 1.5)
            safety_stock = max_usage - avg_usage
            report += f"{name}: Avg Usage/Day {avg_usage:.2f}, Max Usage/Day {max_usage:.2f}, Safety Stock {safety_stock:.2f} {qty_unit}\n"
            safety_stocks.append((name, f"{avg_usage:.2f}", f"{max_usage:.2f}", f"{safety_stock:.2f} {qty_unit}"))

        # Create scrollable window with Treeview
        report_window = tk.Toplevel(self.app.root)
        report_window.title(f"{self.section.capitalize()} Safety Stock Calculation")
        report_window.geometry("600x400")

        tree = ttk.Treeview(report_window, columns=("Name", "Avg Usage/Day", "Max Usage/Day", "Safety Stock"), show="headings")
        tree.heading("Name", text="Name")
        tree.heading("Avg Usage/Day", text="Avg Usage/Day")
        tree.heading("Max Usage/Day", text="Max Usage/Day")
        tree.heading("Safety Stock", text="Safety Stock")
        tree.column("Name", width=150)
        tree.column("Avg Usage/Day", width=100)
        tree.column("Max Usage/Day", width=100)
        tree.column("Safety Stock", width=150)

        for stock in safety_stocks:
            tree.insert("", "end", values=stock)

        scrollbar = ttk.Scrollbar(report_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Button(report_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)
        logging.info("Calculated safety stock")

    # Economic Order Quantity (EOQ)
    def calculate_eoq(self):
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("""
            SELECT i.name, i.quantity, p.price, COUNT(iss.id) as issues 
            FROM items i 
            LEFT JOIN prices p ON i.id = p.item_id 
            LEFT JOIN issuance iss ON i.id = iss.item_id 
            WHERE i.section=? 
            GROUP BY i.id, i.name, i.quantity, p.price
        """, (self.section,))
        items = c.fetchall()
        conn.close()

        order_cost = 10.0  # Adjust as needed
        report = f"{self.section.capitalize()} Economic Order Quantity (EOQ):\n\n"
        eoqs = []
        for name, qty, price, issues in items:
            qty_num, qty_unit = self.app.parse_quantity(qty)
            unit_cost = price if price is not None else 0.0
            annual_demand = issues * 12 if issues else 12
            holding_cost = unit_cost * 0.2
            if holding_cost > 0:
                eoq = ((2 * annual_demand * order_cost) / holding_cost) ** 0.5
                report += f"{name}: Annual Demand {annual_demand}, Unit Cost Ksh{unit_cost:.2f}, EOQ {eoq:.2f} {qty_unit}\n"
                eoqs.append((name, str(annual_demand), f"Ksh{unit_cost:.2f}", f"{eoq:.2f} {qty_unit}"))
            else:
                report += f"{name}: Insufficient cost data for EOQ\n"
                eoqs.append((name, str(annual_demand), "N/A", "N/A"))

        # Create scrollable window with Treeview
        report_window = tk.Toplevel(self.app.root)
        report_window.title(f"{self.section.capitalize()} EOQ Calculation")
        report_window.geometry("600x400")

        tree = ttk.Treeview(report_window, columns=("Name", "Annual Demand", "Unit Cost", "EOQ"), show="headings")
        tree.heading("Name", text="Name")
        tree.heading("Annual Demand", text="Annual Demand")
        tree.heading("Unit Cost", text="Unit Cost")
        tree.heading("EOQ", text="EOQ")
        tree.column("Name", width=150)
        tree.column("Annual Demand", width=100)
        tree.column("Unit Cost", width=100)
        tree.column("EOQ", width=150)

        for eoq in eoqs:
            tree.insert("", "end", values=eoq)

        scrollbar = ttk.Scrollbar(report_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Button(report_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)
        logging.info("Calculated EOQ")

    # Batch Tracking Window
    def manage_batches(self):
        batch_window = tk.Toplevel(self.app.root)
        batch_window.title("Batch Tracking")
        batch_window.geometry("800x600")

        batch_window.columnconfigure(1, weight=1)
        batch_window.rowconfigure(7, weight=1)

        ttk.Label(batch_window, text="Item Name:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.batch_item = ttk.Combobox(batch_window, values=self.get_item_names())
        self.batch_item.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(batch_window, text="Batch Number:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.batch_number = ttk.Entry(batch_window)
        self.batch_number.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(batch_window, text="Quantity (number of purchase units, e.g., '10'):").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.batch_qty = ttk.Entry(batch_window)
        self.batch_qty.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(batch_window, text="Price per Purchase Unit (Ksh):").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.batch_cost = ttk.Entry(batch_window)
        self.batch_cost.grid(row=3, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(batch_window, text="Received Date (YYYY-MM-DD):").grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.batch_received = ttk.Entry(batch_window)
        self.batch_received.grid(row=4, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(batch_window, text="Expiry Date (YYYY-MM-DD):").grid(row=5, column=0, padx=5, pady=5, sticky="w")
        self.batch_expiry = ttk.Entry(batch_window)
        self.batch_expiry.grid(row=5, column=1, padx=5, pady=5, sticky="ew")

        ttk.Button(batch_window, text="Add Batch", command=self.add_batch).grid(row=6, column=1, pady=10, sticky="ew")

        self.batch_tree = ttk.Treeview(batch_window, columns=("ID", "Batch", "Qty", "Cost", "Received", "Expiry"), show="headings")
        self.batch_tree.heading("ID", text="Item ID")
        self.batch_tree.heading("Batch", text="Batch Number")
        self.batch_tree.heading("Qty", text="Quantity")
        self.batch_tree.heading("Cost", text="Price per Purchase Unit (Ksh)")
        self.batch_tree.heading("Received", text="Received Date")
        self.batch_tree.heading("Expiry", text="Expiry Date")
        for col in ("ID", "Batch", "Qty", "Cost", "Received", "Expiry"):
            self.batch_tree.column(col, stretch=tk.YES)
        self.batch_tree.grid(row=7, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")

        self.load_batches()

    def add_batch(self):
        item_name = self.batch_item.get()
        batch_number = self.batch_number.get()
        qty = self.batch_qty.get()  # Number of purchase units
        price_per_unit = float(self.batch_cost.get() or 0.0)  # Price per purchase unit
        received_date = self.batch_received.get()
        expiry_date = self.batch_expiry.get()

        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("SELECT id, quantity, purchase_unit FROM items WHERE section=? AND name=?", (self.section, item_name))
        item = c.fetchone()
        if not item:
            messagebox.showerror("Error", "Item not found")
            conn.close()
            return

        item_id, current_qty, purchase_unit = item
        qty_num = float(qty)
        pu_num, pu_unit = self.app.parse_quantity(purchase_unit)
        total_qty_added = qty_num * pu_num
        curr_num, curr_unit = self.app.parse_quantity(current_qty)
        new_qty = curr_num + total_qty_added
        total_qty_str = f"{new_qty} {curr_unit}"

        # Update price in prices table (price per purchase unit)
        if price_per_unit > 0:
            c.execute("""
                INSERT OR REPLACE INTO prices (item_id, price, last_updated) 
                VALUES (?, ?, ?)
            """, (item_id, price_per_unit, received_date))

        c.execute("""
            INSERT INTO batches (item_id, batch_number, quantity, unit_cost, received_date, expiry_date) 
            VALUES (?, ?, ?, ?, ?, ?)
        """, (item_id, batch_number, f"{qty} x {purchase_unit}", price_per_unit, received_date, expiry_date))
        c.execute("UPDATE items SET quantity=?, last_updated=? WHERE id=?", (total_qty_str, received_date, item_id))
        conn.commit()
        conn.close()

        self.load_batches()
        self.load_items(self.store_tree)
        messagebox.showinfo("Success", f"Batch {batch_number} added for {item_name}")
        logging.info(f"Added batch {batch_number} for {item_name}")
    
    def load_batches(self):
        self.batch_tree.delete(*self.batch_tree.get_children())
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("""
            SELECT b.item_id, b.batch_number, b.quantity, b.unit_cost, b.received_date, b.expiry_date 
            FROM batches b 
            JOIN items i ON b.item_id = i.id 
            WHERE i.section=?
        """, (self.section,))
        for row in c.fetchall():
            self.batch_tree.insert("", "end", values=row)
        conn.close()

    def gmp_analysis(self):
        """
        Good Manufacturing Practices (GMP) Analysis:
        - Checks expiry dates for safety.
        - Ensures sufficient stock and reorder compliance.
        - Verifies batch traceability.
        - Flags items for disposal if expired.
        """
        conn = sqlite3.connect(self.app.get_db_path())
        c = conn.cursor()
        c.execute("""
            SELECT i.name, i.quantity, i.reorder_level, i.expiry_date, i.purchase_unit, 
                   COUNT(b.id) as batch_count, p.price 
            FROM items i 
            LEFT JOIN batches b ON i.id = b.item_id 
            LEFT JOIN prices p ON i.id = p.item_id 
            WHERE i.section=? 
            GROUP BY i.id, i.name, i.quantity, i.reorder_level, i.expiry_date, i.purchase_unit, p.price
        """, (self.section,))
        items = c.fetchall()
        conn.close()

        now = datetime.now()
        report = f"{self.section.capitalize()} GMP Analysis:\n\n"
        expiry_issues = []
        stock_issues = []
        traceability_issues = []
        disposal_items = []

        for name, qty, reorder, expiry_date, purchase_unit, batch_count, price in items:
            qty_num, qty_unit = self.app.parse_quantity(qty)
            reorder_num, _ = self.app.parse_quantity(reorder or "0")  # Default to "0" if None
            pu_num, _ = self.app.parse_quantity(purchase_unit) if purchase_unit else (1, qty_unit)
            num_units = qty_num / pu_num if pu_num and qty_num is not None else 0
            reorder_units = reorder_num / pu_num if pu_num and reorder_num is not None else 0

            # 1. Expiry Check
            if expiry_date:
                try:
                    expiry = datetime.strptime(expiry_date, "%Y-%m-%d")
                    days_left = (expiry - now).days
                    if days_left <= 7:  # Generic threshold
                        expiry_issues.append((name, f"Expires in {days_left} days", expiry_date))
                        if days_left <= 0:
                            disposal_items.append((name, expiry_date, f"{num_units:.2f} x {purchase_unit}"))
                except ValueError:
                    expiry_issues.append((name, "Invalid expiry date", expiry_date))

            # 2. Stock Compliance
            if qty_num is not None and reorder_num is not None and num_units <= reorder_units:
                stock_issues.append((name, f"{num_units:.2f} x {purchase_unit}", f"{reorder_units:.2f} x {purchase_unit}"))

            # 3. Traceability
            if batch_count == 0 and purchase_unit and ("kg" in purchase_unit.lower() or "l" in purchase_unit.lower()):
                traceability_issues.append((name, purchase_unit))

        # Compile Report Sections
        report += "1. Expiry and Safety Check:\n"
        report += "\n".join(f"{item[0]}: {item[1]} ({item[2]})" for item in expiry_issues) + "\n" if expiry_issues else "No expiry issues detected.\n"

        report += "\n2. Stock Compliance:\n"
        report += "\n".join(f"{item[0]}: Qty {item[1]} <= Reorder {item[2]}" for item in stock_issues) + "\n" if stock_issues else "All items above reorder levels.\n"

        report += "\n3. Traceability (Batch Tracking):\n"
        report += "\n".join(f"{item[0]}: No batch records (Bulk item: {item[1]})" for item in traceability_issues) + "\n" if traceability_issues else "Batch tracking adequate for all items.\n"

        report += "\n4. Items for Disposal:\n"
        report += "\n".join(f"{item[0]}: Expired on {item[1]}, Qty: {item[2]}" for item in disposal_items) + "\n" if disposal_items else "No items require disposal.\n"

        # Recommendations
        recommendations = []
        if expiry_issues:
            recommendations.append("Prioritize use or disposal of items nearing expiry.")
        if stock_issues:
            recommendations.append("Order additional stock for items below reorder levels.")
        if traceability_issues:
            recommendations.append("Implement batch tracking for bulk items.")
        if disposal_items:
            recommendations.append("Dispose of expired items per sanitation guidelines.")
        report += "\nRecommendations:\n"
        report += "\n".join(recommendations) + "\n" if recommendations else "Maintain current practices.\n"

        # Create scrollable window with Treeview
        report_window = tk.Toplevel(self.app.root)
        report_window.title(f"{self.section.capitalize()} GMP Analysis")
        report_window.geometry("800x500")

        tree = ttk.Treeview(report_window, columns=("Category", "Item", "Details"), show="headings")
        tree.heading("Category", text="Category")
        tree.heading("Item", text="Item")
        tree.heading("Details", text="Details")
        tree.column("Category", width=200)
        tree.column("Item", width=200)
        tree.column("Details", width=350)

        # Populate Treeview
        if expiry_issues:
            tree.insert("", "end", values=("Expiry Issues", "", ""))
            for item in expiry_issues:
                tree.insert("", "end", values=("", item[0], f"{item[1]} ({item[2]})"))
        else:
            tree.insert("", "end", values=("Expiry Issues", "No issues", ""))

        if stock_issues:
            tree.insert("", "end", values=("Stock Compliance", "", ""))
            for item in stock_issues:
                tree.insert("", "end", values=("", item[0], f"Qty {item[1]} <= Reorder {item[2]}"))
        else:
            tree.insert("", "end", values=("Stock Compliance", "All adequate", ""))

        if traceability_issues:
            tree.insert("", "end", values=("Traceability", "", ""))
            for item in traceability_issues:
                tree.insert("", "end", values=("", item[0], f"No batch records (Bulk item: {item[1]})"))
        else:
            tree.insert("", "end", values=("Traceability", "Adequate", ""))

        if disposal_items:
            tree.insert("", "end", values=("Disposal Items", "", ""))
            for item in disposal_items:
                tree.insert("", "end", values=("", item[0], f"Expired on {item[1]}, Qty: {item[2]}"))
        else:
            tree.insert("", "end", values=("Disposal Items", "None required", ""))

        if recommendations:
            tree.insert("", "end", values=("Recommendations", "", ""))
            for rec in recommendations:
                tree.insert("", "end", values=("", "", rec))
        else:
            tree.insert("", "end", values=("Recommendations", "Maintain current practices", ""))

        scrollbar = ttk.Scrollbar(report_window, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Button(report_window, text="Download PDF", command=lambda: self.download_audit_pdf(report)).pack(pady=5)
        logging.info("Performed GMP analysis")